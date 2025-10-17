# DragonLog (c) 2023-2025 by Andreas Schawo is licensed under CC BY-SA 4.0.
# To view a copy of this license, visit http://creativecommons.org/licenses/by-sa/4.0/

import os
import csv
import sys
import json
import string
import zipfile
import platform
from enum import Enum, auto
import datetime
from typing import Iterable, Iterator

from packaging.version import parse, Version, InvalidVersion
from PyQt6 import QtCore, QtWidgets, QtSql, QtGui
import adif_file
from adif_file import adi, adx
import xmltodict
import hamcc
import requests

try:
    # noinspection PyPackageRequirements,PyUnresolvedReferences,PyProtectedMember
    from cv2 import __version__ as cv2_version
    # noinspection PyUnresolvedReferences
    from pyzbar import __version__ as pyzbar_version

    OPTION_QRCODEREADER = True
except ImportError:
    OPTION_QRCODEREADER = False

try:
    # noinspection PyUnresolvedReferences
    import openpyxl
    # noinspection PyUnresolvedReferences
    from openpyxl.styles import Font

    OPTION_OPENPYXL = True
except ImportError:
    OPTION_OPENPYXL = False

from . import DragonLog_MainWindow_ui
from .Logger import Logger
from .RegEx import find_non_ascii, check_format, REGEX_DATE
from .DragonLog_QSOForm import QSOForm
from .DragonLog_Settings import Settings
from .DragonLog_AppSelect import AppSelect
from .LoTW import (LoTW, LoTWADIFFieldException, LoTWRequestException, LoTWCommunicationException,
                   LoTWLoginException, LoTWNoRecordException)
from .eQSL import (EQSL, EQSLADIFFieldException, EQSLLoginException, EQSLRequestException,
                   EQSLUserCallMatchException, EQSLQSODuplicateException, EQSLCommunicationException)
from .CassiopeiaConsole import CassiopeiaConsole
from .CallBook import HamQTHCallBook, CallBookType, LoginException, QSORejectedException, MissingADIFFieldException, \
    CommunicationException
from .DxSpots import DxSpots
from .ContestDlg import ContestDialog
from .ContestStatistics import ContestStatistics
from .contest import CONTESTS, CONTEST_IDS, CONTEST_NAMES, build_contest_list, ExchangeData
from .distance import distance
from .cty import CountryData, Country, CountryNotFoundException, CountryCodeNotFoundException
from .RigControl import RigControl, RigctldNotConfiguredException, RigctldExecutionException, \
    CATSettingsMissingException
from . import ColorPalettes
from .DragonLog_Statistics import StatisticsWidget
from .local_callbook import (LocalCallbook, LocalCallbookData, CallHistoryData,
                             LocalCallbookDatabaseError, LocalCallbookImportError, LocalCallbookExportError)

if OPTION_QRCODEREADER:
    from .QSLQRReader import QSLQRReaderDialog

__prog_name__ = 'DragonLog'
__prog_desc__ = 'Log QSO for Ham radio'
__author_name__ = 'Andreas Schawo, DF1ASC'
__author_email__ = 'andreas@schawo.de'
__copyright__ = 'Copyright 2023-2025 by Andreas Schawo,licensed under CC BY-SA 4.0'

from . import __version__ as version

__version__ = version.__version__
__version_str__ = version.__version_str__

if version.__branch__ != 'master':
    __version_str__ += '-' + version.__branch__
if version.__unclean__:
    __version_str__ += '-unclean'


def update_available(cur_version: str) -> tuple[bool, str]:
    try:
        cur_ver: Version = parse(cur_version)
        req = requests.get('https://pypi.python.org/pypi/dragonlog/json')
        if req.status_code == requests.codes.ok:
            pypi_info = req.json()
            if pypi_info and 'info' in pypi_info:
                pypi_ver: Version = parse(pypi_info['info'].get('version', '0'))
                return (not pypi_ver.is_prerelease and cur_ver < pypi_ver), str(pypi_ver)
    except (InvalidVersion, requests.exceptions.ConnectionError):
        pass

    return False, ''


class DatabaseOpenException(Exception):
    pass


class DatabaseWriteException(Exception):
    pass


class BackgroundBrushDelegate(QtWidgets.QStyledItemDelegate):
    """A delegate to change background color depending on a columns content and translation of different columns"""

    def __init__(self, color_map: dict, column: int):
        super(BackgroundBrushDelegate, self).__init__()

        self.color_map: dict = color_map
        self.column: int = column

    def getColor(self, value):
        color = [255, 255, 255, 0]  # white as fallback

        if value in self.color_map:
            color = self.color_map[value]
        elif 'default' in self.color_map:
            color = self.color_map['default']

        return color

    def initStyleOption(self, option: QtWidgets.QStyleOptionViewItem, index: QtCore.QModelIndex):
        super().initStyleOption(option, index)

        option.backgroundBrush = QtGui.QBrush(
            QtGui.QColor(*self.getColor(index.model().data(index.siblingAtColumn(self.column)))))


class TranslatedTableModel(QtSql.QSqlTableModel):
    """Translate propagation values and status to clear text and fancy icon for status"""

    def __init__(self, parent, db_conn, status_cols: Iterable, cols: tuple, prop_tr: dict):
        super(TranslatedTableModel, self).__init__(parent, db_conn)

        self.status_cols = status_cols
        self.status_translation = {
            'Y': self.tr('Y'),
            'N': self.tr('N'),
            'M': self.tr('M'),
            'R': self.tr('R'),
        }

        self.prop_col = cols.index('propagation')
        self.prop_translation = prop_tr

        self.freq_col = cols.index('freq')
        self.pwr_col = cols.index('power')
        self.dist_col = cols.index('dist')
        self.contest_col = cols.index('contest_id')
        self.call_cols = cols.index('own_callsign'), cols.index('call_sign')
        self.loc_cols = cols.index('own_locator'), cols.index('locator')

        # noinspection PyUnresolvedReferences
        self.ok_icon = QtGui.QIcon(self.parent().searchFile('icons:ok.png'))
        # noinspection PyUnresolvedReferences
        self.no_icon = QtGui.QIcon(self.parent().searchFile('icons:no.png'))

    def data(self, idx, role=QtCore.Qt.ItemDataRole.DisplayRole):
        value = super().data(idx, role)
        if role == QtCore.Qt.ItemDataRole.DisplayRole:
            if idx.column() in self.status_cols and value in self.status_translation:
                return self.status_translation[value]
            elif idx.column() == self.prop_col and value in self.prop_translation:
                return self.prop_translation[value]
            elif idx.column() == self.freq_col and value:
                return f'{value:.3f} kHz'
            elif idx.column() == self.pwr_col and value:
                return f'{int(value)} W'
            elif idx.column() == self.dist_col and value:
                return f'{int(value)} km'
            elif idx.column() == self.contest_col and value in CONTEST_NAMES:
                return CONTEST_NAMES[value]

        if role == QtCore.Qt.ItemDataRole.DecorationRole:
            txt = super().data(idx, QtCore.Qt.ItemDataRole.DisplayRole)
            if idx.column() in self.status_cols:
                if txt in ('Y', 'M'):
                    return self.ok_icon
                else:
                    return self.no_icon

        return value


class ADISourceType(Enum):
    Other = auto()
    HamQTH = auto()
    QRZCQ = auto()
    QRZ = auto()
    eQSLInbox = auto()
    eQSLOutbox = auto()
    LoTW = auto()
    DCL = auto()


__adi_src_map__ = {
    'eQSL.cc DownloadInBox': ADISourceType.eQSLInbox,
    'eQSL.cc DownloadADIF': ADISourceType.eQSLOutbox,
    'LoTW': ADISourceType.LoTW,
    'QRZLogbook': ADISourceType.QRZ,
    'DCL': ADISourceType.DCL,
}


def eval_adi_type(prog_id: str) -> ADISourceType:
    if prog_id in __adi_src_map__:
        return __adi_src_map__[prog_id]
    else:
        return ADISourceType.Other


# Translate ADIF contests IDs to internal and vv
CONTEST_ADIF_DL = {
    'DARC-UKW-SPRING': 'DARC-UKW-FRUEHLING',
    'DARC-UKW-FIELD-DAY': 'DARC-UKW-SOMMERFD',
    'DARC-MICROWAVE': 'DARC-UKW',
    'EASTER': 'DARC-KW-OSTERN',
}
CONTEST_DL_ADIF = dict([(v, k) for k, v in CONTEST_ADIF_DL.items()])


# noinspection PyPep8Naming
class DragonLog(QtWidgets.QMainWindow, DragonLog_MainWindow_ui.Ui_MainWindow):
    qsoAdded = QtCore.pyqtSignal()
    qsoChanged = QtCore.pyqtSignal()
    qsoDeleted = QtCore.pyqtSignal()

    __sql_cols__ = (
        'id', 'date_time', 'date_time_off', 'own_callsign', 'call_sign', 'name', 'qth', 'locator',
        'rst_sent', 'rst_rcvd', 'band', 'mode', 'submode', 'freq', 'channel', 'power', 'propagation',
        'own_name', 'own_qth', 'own_locator', 'radio', 'antenna',
        'remarks', 'comments', 'dist',
        'qsl_via', 'qsl_path', 'qsl_msg', 'qsl_sent', 'qsl_rcvd',
        'eqsl_sent', 'eqsl_rcvd', 'lotw_sent', 'lotw_rcvd', 'hamqth',
        'contest_id', 'ctx_qso_id', 'crx_qso_id', 'crx_data', 'event', 'evt_tx_exch', 'evt_rx_exch',
    )

    __adx_cols__ = (
        'QSO_DATE/TIME_ON', 'QSO_DATE/TIME_OFF', 'STATION_CALLSIGN', 'CALL', 'NAME_INTL', 'QTH_INTL', 'GRIDSQUARE',
        'RST_SENT', 'RST_RCVD', 'BAND', 'MODE', 'SUBMODE', 'FREQ', 'APP_DRAGONLOG_CBCHANNEL', 'TX_PWR', 'PROP_MODE',
        'MY_NAME_INTL', 'MY_CITY_INTL', 'MY_GRIDSQUARE', 'MY_RIG_INTL', 'MY_ANTENNA_INTL',
        'NOTES_INTL', 'COMMENT_INTL', 'DISTANCE',
        'QSL_VIA', 'QSL_SENT_VIA', 'QSLMSG_INTL', 'QSL_SENT', 'QSL_RCVD',
        'EQSL_QSL_SENT', 'EQSL_QSL_RCVD', 'LOTW_QSL_SENT', 'LOTW_QSL_RCVD', 'HAMQTH_QSO_UPLOAD_STATUS',
        'CONTEST_ID', 'STX', 'SRX', 'SRX_STRING', 'MY_SIG_INTL', 'MY_SIG_INFO_INTL', 'SIG_INFO_INTL',
    )

    __db_create_stmnt__ = '''CREATE TABLE IF NOT EXISTS "qsos" (
                            "id"    INTEGER PRIMARY KEY NOT NULL,
                            "date_time"   NUMERIC,
                            "date_time_off"   NUMERIC,
                            "own_callsign" TEXT,
                            "call_sign"  TEXT,
                            "name"  TEXT,
                            "qth"    TEXT,
                            "locator" TEXT,
                            "rst_sent" TEXT,
                            "rst_rcvd" TEXT,
                            "band"    TEXT,
                            "mode"   TEXT,
                            "submode"  TEXT,
                            "freq"  REAL,
                            "channel"  INTEGER,
                            "power"  REAL,
                            "propagation"  TEXT,
                            "own_name"  TEXT,
                            "own_qth"   TEXT,
                            "own_locator" TEXT,
                            "radio"   TEXT,
                            "antenna"   TEXT,
                            "remarks"   TEXT,
                            "comments"   TEXT,
                            "dist" INTEGER,
                            "qsl_via"   TEXT,
                            "qsl_path"   TEXT,
                            "qsl_msg"   TEXT,
                            "qsl_sent"   TEXT,
                            "qsl_rcvd"   TEXT,
                            "eqsl_sent"   TEXT,
                            "eqsl_rcvd"   TEXT,
                            "lotw_sent"   TEXT,
                            "lotw_rcvd"   TEXT,
                            "hamqth"   TEXT,
                            "contest_id" TEXT,
                            "ctx_qso_id" INTEGER,
                            "crx_qso_id" INTEGER,
                            "crx_data" TEXT,
                            "event" TEXT, 
                            "evt_tx_exch" TEXT, 
                            "evt_rx_exch" TEXT
                        );'''

    __db_create_idx_stmnt__ = '''CREATE INDEX IF NOT EXISTS "find_qso" ON "qsos" (
                                    "date_time",
                                    "call_sign"
                                )'''

    __db_insert_stmnt__ = 'INSERT INTO qsos (' + ",".join(__sql_cols__[1:]) + ') ' \
                                                                              'VALUES (' + ",".join(
        ["?"] * (len(__sql_cols__) - 1)) + ')'
    __db_update_stmnt__ = 'UPDATE qsos SET ' + "=?,".join(__sql_cols__[1:]) + '=? ' \
                                                                              'WHERE id = ?'
    __db_select_stmnt__ = 'SELECT * FROM qsos'

    __db_create_view_stmnt_qso_count__ = '''CREATE VIEW IF NOT EXISTS qso_count AS 
                                      SELECT COUNT(id) as qsos FROM qsos;'''

    @staticmethod
    def searchFile(name) -> str | None:
        file = QtCore.QFile(name)
        if file.exists():
            return file.fileName()
        else:
            return None

    def __init__(self, file=None, app_path='.', ini_file=''):
        super().__init__()
        self.setupUi(self)

        self.app_path = app_path
        self.help_dialog = None
        self.help_sc_dialog = None
        self.help_cc_dialog = None
        self.help_contest_dialog = None
        self.qso_form = None
        self.dxspots_widget = None
        self.cstats_widget = None

        # Database
        self.__db_con__ = QtSql.QSqlDatabase.addDatabase('QSQLITE', 'main')

        if ini_file:
            self.settings = QtCore.QSettings(ini_file, QtCore.QSettings.Format.IniFormat)
        else:
            self.settings = QtCore.QSettings(self.programName)

        self.log = Logger(self.logTextEdit, self.settings)
        self.log.info(f'Starting {self.programName} {self.programVersion}...')
        self.log.debug(f'Platform: {platform.platform()}')
        self.log.debug(f'Python: {sys.version}')
        if not OPTION_OPENPYXL:
            self.log.info(f'Option XL-Format not available')
        if not OPTION_QRCODEREADER:
            self.log.info(f'Option QSL-QR-Code not available')
        self.log.info(f'Using settings {self.settings.format()} from "{self.settings.fileName()}"')

        try:
            font_inter_id = QtGui.QFontDatabase.addApplicationFont(self.searchFile('data:InterSlashedZero.ttf'))
            font_name = QtGui.QFontDatabase.applicationFontFamilies(font_inter_id)[0]
            self.log.debug(f'Registered font: "{font_name}"')
        except IndexError:
            self.log.warning('Could not register special font')

        if int(self.settings.value('ui/log_dock_float', 0)):
            self.logDockWidget.setFloating(True)
        else:
            log_dock_area = self.int2dock_area(
                int(self.settings.value('ui/log_dock_area',
                                        QtCore.Qt.DockWidgetArea.BottomDockWidgetArea.value)))
            self.addDockWidget(log_dock_area, self.logDockWidget)
        self.logDockWidget.setVisible(bool(int(self.settings.value('ui/show_log', 0))))

        if int(self.settings.value('ui/filter_dock_float', 0)):
            self.filterDockWidget.setFloating(True)
            self.filterDockWidget.resize(10, 10)
        else:
            filter_dock_area = self.int2dock_area(
                int(self.settings.value('ui/filter_dock_area',
                                        QtCore.Qt.DockWidgetArea.TopDockWidgetArea.value)))
            self.addDockWidget(filter_dock_area, self.filterDockWidget)
        self.filterDockWidget.setVisible(bool(int(self.settings.value('ui/show_filter', 0))))

        self.__recent_filter__ = ''
        self.__table_filter__ = ''
        self.filterDockWidget.visibilityChanged.connect(self.resetTableFilter)
        self.filterDockWidget.dockLocationChanged.connect(self.filterWidgetResize)
        self.refreshFilterStatus()

        self.callbook_status = QtWidgets.QLabel(f'{self.tr("Callbook")}: {self.tr("None")}')
        self.statusBar().addPermanentWidget(self.callbook_status)
        self.dummy_status = QtWidgets.QLabel()
        self.statusBar().addPermanentWidget(self.dummy_status, 1)
        self.watch_status = QtWidgets.QLabel(self.tr('Watching file') + ': ' + self.tr('inactiv'))
        self.statusBar().addPermanentWidget(self.watch_status)
        self.hamlib_status = QtWidgets.QLabel(self.tr('Hamlib') + ': ' + self.tr('inactiv'))
        self.statusBar().addPermanentWidget(self.hamlib_status)
        self.hamlib_error = QtWidgets.QLabel('')
        self.statusBar().addPermanentWidget(self.hamlib_error)

        with open(self.searchFile('data:bands.json')) as bj:
            self.bands: dict = json.load(bj)
        with open(self.searchFile('data:modes.json')) as mj:
            self.modes: dict = json.load(mj)
        with open(self.searchFile('data:cb_channels.json')) as cj:
            self.cb_channels: dict = json.load(cj)
        with open(self.searchFile('data:color_map.json')) as cmj:
            color_map: dict = json.load(cmj)
        self.QSOTableView.setItemDelegate(BackgroundBrushDelegate(color_map, self.__sql_cols__.index('band')))

        self.fBandComboBox.insertItem(0, '')
        self.fBandComboBox.insertItems(1, self.bands.keys())
        self.fModeComboBox.insertItem(0, '')
        self.fModeComboBox.insertItems(1, self.modes['AFU'].keys())
        self.fContestComboBox.insertItem(0, '')
        self.fContestComboBox.insertItems(1, CONTEST_NAMES.values())

        self.prop: dict = {
            '': '',
            'AS': self.tr('Aircraft Scatter'),
            'AUE': self.tr('Aurora-E'),
            'AUR': self.tr('Aurora'),
            'BS': self.tr('Back scatter'),
            'ECH': self.tr('EchoLink'),
            'EME': self.tr('Earth-Moon-Earth'),
            'ES': self.tr('Sporadic E'),
            'F2': self.tr('F2 Reflection'),
            'FAI': self.tr('Field Aligned Irregularities'),
            'GWAVE': self.tr('Ground Wave'),
            'INTERNET': self.tr('Internet-assisted'),
            'ION': self.tr('Ionoscatter'),
            'IRL': self.tr('IRLP'),
            'LOS': self.tr('Line of Sight'),
            'MS': self.tr('Meteor scatter'),
            'RPT': self.tr('Repeater or Transponder'),
            'RS': self.tr('Rain scatter'),
            'SAT': self.tr('Satellite'),
            'TEP': self.tr('Trans-equatorial'),
            'TR': self.tr('Tropospheric ducting'),
        }

        self.ascii_replace: dict = {}
        # noinspection PyBroadException
        try:
            with open(self.searchFile(f'data:i18n/ascii_replace_{QtCore.QLocale.system().name()[:2]}.json'),
                      encoding='utf-8') as arf:
                self.ascii_replace = json.load(arf)
                self.log.debug(f'Loaded ascii_replace from "{arf.name}"')
        except Exception:
            self.log.debug(f'Found no ascii_replace_xx.json for "{print(QtCore.QLocale.system().name()[:2])}"')

        self.__headers__ = (
            self.tr('QSO'),
            self.tr('Date/Time start'),
            self.tr('Date/Time end'),
            self.tr('Own call sign'),
            self.tr('Call sign'),
            self.tr('Name'),
            self.tr('QTH'),
            self.tr('Locator'),
            self.tr('RST sent'),
            self.tr('RST rcvd'),
            self.tr('Band'),
            self.tr('Mode'),
            self.tr('Submode'),
            self.tr('Frequency'),
            self.tr('Channel'),
            self.tr('Power'),
            self.tr('Propagation'),
            self.tr('Own Name'),
            self.tr('Own QTH'),
            self.tr('Own Locator'),
            self.tr('Radio'),
            self.tr('Antenna'),
            self.tr('Notes'),
            self.tr('Comment'),
            self.tr('Distance'),
            self.tr('QSL via'),
            self.tr('QSL path'),
            self.tr('QSL message'),
            self.tr('QSL sent'),
            self.tr('QSL rcvd'),
            self.tr('eQSL sent'),
            self.tr('eQSL rcvd'),
            self.tr('LoTW sent'),
            self.tr('LoTW rcvd'),
            self.tr('HamQTH'),
            self.tr('Contest'),
            self.tr('Tx QSO ID'),
            self.tr('Rx QSO ID'),
            self.tr('Rx data'),
            self.tr('Event'),
            self.tr('Tx Exch'),
            self.tr('Rx Exch'),
        )

        self.__header_map__ = dict(zip(self.__sql_cols__, self.__headers__))

        self.__rigctl__ = RigControl(self, self.settings, self.log, self.bands, self.modes)
        self.__rigctl__.statusChanged.connect(self.rigStatusChanged)

        self.__cty__: CountryData | None = None
        self.cty_load(self.settings.value('dx_spots/cty_data', ''))

        # Settings Form
        self.settings_form = Settings(self, self.settings, self.bands, self.modes['AFU'],
                                      self.__headers__, self.log, self.__rigctl__)
        self.settings_form.ctyDataChanged.connect(self.cty_load)

        self.__local_cb__ = None
        if self.settings.value('lastCallbookPath', None):
            try:
                self.__local_cb__ = LocalCallbook(self.settings.value('lastCallbookPath'), self.log,
                                                  self.settings_form.csv_delimiter)
                self.updateCallbookStatus()
                self.actionRefresh_callbook.setEnabled(True)
                self.actionExport_callbook.setEnabled(True)
                self.actionImport_callbook.setEnabled(True)
                self.actionRefresh_call_history.setEnabled(True)
                self.actionExport_call_history.setEnabled(True)
                self.actionImport_call_history.setEnabled(True)
            except LocalCallbookDatabaseError as exc:
                self.log.exception(exc)

        # QSOForm
        self.qso_form = QSOForm(self, self, self.bands, self.modes, self.prop, self.settings, self.settings_form,
                                self.cb_channels, self.log, self.__local_cb__)
        self.qsoDockWidget.setWidget(self.qso_form)
        self.qsoDockWidget.visibilityChanged.connect(self.qso_form.startTimers)
        self.qsoDockWidget.visibilityChanged.connect(self.qso_form.clear)

        if int(self.settings.value('ui/qso_dock_float', 0)):
            self.qsoDockWidget.setFloating(True)
        else:
            qso_dock_area = self.int2dock_area(
                int(self.settings.value('ui/qso_dock_area',
                                        QtCore.Qt.DockWidgetArea.RightDockWidgetArea.value)))
            self.addDockWidget(qso_dock_area,
                               self.qsoDockWidget)
        self.qsoDockWidget.setVisible(bool(int(self.settings.value('ui/show_qso', 0))))
        self.__rigctl__.bandChanged.connect(self.qso_form.setBand)
        self.__rigctl__.frequencyChanged.connect(self.qso_form.setFrequency)
        self.__rigctl__.modeChanged.connect(self.qso_form.setMode)
        self.__rigctl__.submodeChanged.connect(self.qso_form.setSubmode)
        self.__rigctl__.powerChanged.connect(self.qso_form.setPower)

        # HamCCForm
        self.cc_widget = CassiopeiaConsole(self, self, self.settings, self.log, self.__local_cb__)
        self.cc_widget.qsosChached.connect(self.fetchCCQSO)
        self.ccDockWidget.setWidget(self.cc_widget)
        if int(self.settings.value('ui/cc_dock_float', 0)):
            self.ccDockWidget.setFloating(True)
        else:
            cc_dock_area = self.int2dock_area(
                int(self.settings.value('ui/cc_dock_area',
                                        QtCore.Qt.DockWidgetArea.BottomDockWidgetArea.value)))
            self.addDockWidget(cc_dock_area,
                               self.ccDockWidget)
        self.ccDockWidget.setVisible(bool(int(self.settings.value('ui/show_cc', 0))))
        self.__rigctl__.frequencyChanged.connect(self.cc_widget.setFrequency)
        self.__rigctl__.bandChanged.connect(self.cc_widget.setBand)
        self.__rigctl__.modeChanged.connect(self.cc_widget.setMode)
        self.__rigctl__.powerChanged.connect(self.cc_widget.setPower)

        # DxSpotsForm
        self.dxspots_widget = DxSpots(self, self, self.settings, self.log)
        self.dxSpotsDockWidget.setWidget(self.dxspots_widget)
        if int(self.settings.value('ui/dxspots_dock_float', 0)):
            self.dxSpotsDockWidget.setFloating(True)
        else:
            dxspots_dock_area = self.int2dock_area(
                int(self.settings.value('ui/dxspots_dock_area',
                                        QtCore.Qt.DockWidgetArea.LeftDockWidgetArea.value)))
            self.addDockWidget(dxspots_dock_area,
                               self.dxSpotsDockWidget)
        self.dxSpotsDockWidget.setVisible(bool(int(self.settings.value('ui/show_dxspots', 0))))
        self.dxspots_widget.spotSelected.connect(self.setQSO)

        # ContestStatistics
        self.cstats_widget = ContestStatistics(self, self, self.settings, self.log, self.__cty__)
        self.contestStatDockWidget.setWidget(self.cstats_widget)
        if int(self.settings.value('ui/contest_stats_dock_float', 0)):
            self.dxSpotsDockWidget.setFloating(True)
        else:
            contest_stats_dock_area = self.int2dock_area(
                int(self.settings.value('ui/contest_stats_dock_area',
                                        QtCore.Qt.DockWidgetArea.BottomDockWidgetArea.value)))
            self.addDockWidget(contest_stats_dock_area,
                               self.contestStatDockWidget)
        self.contestStatDockWidget.setVisible(bool(int(self.settings.value('ui/show_contest_stats', 0))))
        self.qsoAdded.connect(self.cstats_widget.fetchQSOs)
        self.qsoChanged.connect(self.cstats_widget.fetchQSOs)
        self.qsoDeleted.connect(self.cstats_widget.fetchQSOs)
        self.cstats_widget.contestSelected.connect(self.fContestComboBox.setCurrentText)
        self.cstats_widget.contestSelected.connect(self.cc_widget.eventChanged)
        self.cstats_widget.fromDateSelected.connect(self.fDateFromLineEdit.setText)
        self.cstats_widget.fromDateSelected.connect(self.fDateFromLineEdit.editingFinished)
        self.cstats_widget.toDateSelected.connect(self.fDateToLineEdit.setText)
        self.cstats_widget.toDateSelected.connect(self.fDateToLineEdit.editingFinished)

        self.keep_logging = False

        if file:
            self.log.info(f'Opening database from commandline {file}...')
            self.connectDB(file)
        elif self.settings.value('lastDatabase', None):
            if os.path.isfile(self.settings.value('lastDatabase', None)):
                self.log.info(f'Opening last database {self.settings.value("lastDatabase", None)}...')
                self.connectDB(self.settings.value('lastDatabase', None))
            else:
                self.log.warning(f'Opening last database {self.settings.value("lastDatabase", None)} failed!')

        self.watchAppSelect = AppSelect(self, f'{self.programName} - {self.tr("Watch application log")}',
                                        self.settings)
        self.watchTimer = QtCore.QTimer(self)
        self.watchTimer.timeout.connect(self.watchFile)
        self.watchPos = 0
        self.watchFileName = ''

        self.lotw = LoTW(self.log)

        self.settings_form.settingsStored.connect(self.qso_form.refreshBands)
        self.settings_form.settingsStored.connect(self.qso_form.refreshQTHList)
        self.settings_form.settingsStored.connect(self.qso_form.refreshRadioList)
        self.settings_form.settingsStored.connect(self.qso_form.refreshAntennaList)
        self.settings_form.settingsStored.connect(self.qso_form.refreshOwnData)
        self.settings_form.settingsStored.connect(self.cc_widget.resetUserData)
        self.settings_form.settingsStored.connect(self.cc_widget.refreshListings)

        if 'station' not in self.settings.childGroups():
            QtWidgets.QMessageBox.information(self, self.tr('Initial startup'),
                                              self.tr('It seems you are running DragonLog for the first time\n'
                                                      'Please configure some information first'))
            self.showSettings()

        self.eqsl = EQSL(self.programName, self.log)
        self.eqsl_urls: dict[str, str] = {}

        self.useFont()

        if (self.settings.value('ui/check_updates', 1) and
                not self.settings.value('ui/update_checked', '') == __version__):
            upd, upd_ver = update_available(__version__)
            self.log.debug(f'Checked update: {upd}, new version "{upd_ver}"')
            if upd:
                QtWidgets.QMessageBox.information(self, self.tr('Update available'),
                                                  self.tr('A newer version of DragonLog is available') + f': {upd_ver}')
                self.settings.setValue('ui/update_checked', __version__)

    @staticmethod
    def int2dock_area(value: int) -> QtCore.Qt.DockWidgetArea:
        dock_area = QtCore.Qt.DockWidgetArea.NoDockWidgetArea
        match value:
            case 1:
                dock_area = QtCore.Qt.DockWidgetArea.LeftDockWidgetArea
            case 2:
                dock_area = QtCore.Qt.DockWidgetArea.RightDockWidgetArea
            case 4:
                dock_area = QtCore.Qt.DockWidgetArea.TopDockWidgetArea
            case 8:
                dock_area = QtCore.Qt.DockWidgetArea.BottomDockWidgetArea

        return dock_area

    def setQSO(self, call: str, band: str, freq: float):
        if self.__rigctl__.isActive():
            self.__rigctl__.setRigFreq(freq)
            self.qso_form.setQSO(call)
            self.cc_widget.setQSO(call)
        else:
            self.qso_form.setQSO(call, band, freq)
            self.cc_widget.setQSO(call, band, freq)

    def filterWidgetResize(self):
        if self.filterDockWidget.isFloating():
            self.filterDockWidget.resize(10, 10)

    def showSettings(self):
        if self.settings_form.exec():
            self.refreshTableView()
            self.useFont()

    def selectDB(self):
        res = QtWidgets.QFileDialog.getSaveFileName(
            self,
            self.tr('Select file'),
            self.settings.value('lastDatabase', None),
            self.tr('QSO-Log (*.qlog);;All Files (*.*)'),
            options=QtWidgets.QFileDialog.Option.DontConfirmOverwrite)

        if res[0]:
            self.log.info(f'Selected database {res[0]}')
            self.connectDB(res[0])

    def checkDB(self, db_file):
        # Check database for missing cols
        res = self.__db_con__.exec('SELECT GROUP_CONCAT(NAME,",") as columns FROM PRAGMA_TABLE_INFO("qsos")')
        res.next()
        db_cols = res.value('columns')
        if not db_cols:
            QtWidgets.QMessageBox.critical(self, self.tr('Database error'),
                                           self.tr('Checking database failed. Content is not accessable.'))
            self.close()
            return False

        db_cols_l = db_cols.split(',')
        missing_cols = False
        for col in self.__sql_cols__:
            if col not in db_cols_l:
                missing_cols = True
                break

        if missing_cols:
            path, db_name = os.path.split(db_file)
            bck_name = os.path.join(path, datetime.date.today().strftime('%Y-%m-%d') + ' ' + db_name)

            QtWidgets.QMessageBox.warning(self, self.tr('Database structure out-dated'),
                                          self.tr('The database structure is out-dated and needs a conversion\n\n'
                                                  'A backup will be generated:') + f'\n"{bck_name}"',
                                          )

            # Create backup
            self.__db_con__.close()
            try:
                os.rename(db_file, bck_name)
                self.log.info(f'Database backup at "{bck_name}"')
            except FileExistsError:
                self.log.error('A database backup could not be created. The file already exists.')
                QtWidgets.QMessageBox.critical(self, self.tr('Database backup error'),
                                               self.tr('A database backup could not be created.\n'
                                                       'The file already exists.'))
                self.close()
                return False

            # Open new DB
            self.__db_con__.setDatabaseName(db_file)
            if self.__db_con__.lastError().text():
                raise DatabaseOpenException(self.__db_con__.lastError().text())
            self.__db_con__.open()
            self.__db_con__.exec('PRAGMA user_version = 0xDF1A5C;')
            self.__db_con__.exec('PRAGMA application_id = 0x106b004;')
            self.__db_con__.exec('PRAGMA temp_store = MEMORY;')
            self.__db_con__.exec('PRAGMA journal_mode = WAL;')
            self.__db_con__.exec('PRAGMA synchronous = NORMAL;')

            self.__db_con__.exec(self.__db_create_stmnt__)
            self.__db_con__.exec(self.__db_create_idx_stmnt__)
            self.__db_con__.exec(self.__db_create_view_stmnt_qso_count__)

            if self.__db_con__.lastError().text():
                raise DatabaseOpenException(self.__db_con__.lastError().text())

            self.log.debug('Initialised new database for conversion')

            # Open backup
            db_con_bck = QtSql.QSqlDatabase.addDatabase('QSQLITE', 'backup')
            db_con_bck.setDatabaseName(str(bck_name))
            if db_con_bck.lastError().text():
                raise DatabaseOpenException(db_con_bck.lastError().text())
            db_con_bck.open()
            self.__db_con__.exec(self.__db_create_stmnt__)
            if self.__db_con__.lastError().text():
                raise DatabaseOpenException(self.__db_con__.lastError().text())

            self.log.debug('Opened DB backup for conversion')

            q_read = db_con_bck.exec(f'SELECT {db_cols} FROM qsos')
            if q_read.lastError().text():
                raise Exception(q_read.lastError().text())

            insert_stmnt = f'INSERT INTO qsos ({db_cols}) ' \
                           'VALUES (' + ",".join(["?"] * len(db_cols_l)) + ')'

            while q_read.next():
                row = []
                for i in range(len(db_cols_l)):
                    row.append(q_read.value(i))

                q_write = QtSql.QSqlQuery(self.__db_con__)
                q_write.prepare(insert_stmnt)
                for i, val in enumerate(row):
                    q_write.bindValue(i, val)
                q_write.exec()
                if q_write.lastError().text():
                    raise Exception(q_write.lastError().text())

            self.__db_con__.commit()
            db_con_bck.close()

            self.log.info('Database conversion finished successfully')
            QtWidgets.QMessageBox.information(self, self.tr('Database conversion'),
                                              self.tr('Database conversion finished')
                                              )

        return True

    def connectDB(self, db_file):
        db_file = os.path.abspath(db_file)
        try:
            if self.__db_con__.isOpen():
                self.__db_con__.close()

            self.__db_con__.setDatabaseName(db_file)
            self.__db_con__.open()
            self.__db_con__.exec('PRAGMA user_version = 0xDF1A5C;')
            self.__db_con__.exec('PRAGMA application_id = 0x106b004;')
            self.__db_con__.exec('PRAGMA temp_store = MEMORY;')
            self.__db_con__.exec('PRAGMA journal_mode = WAL;')
            self.__db_con__.exec('PRAGMA synchronous = NORMAL;')

            self.log.debug('Initialise database if necessary...')
            self.__db_con__.exec(self.__db_create_stmnt__)
            self.__db_con__.exec(self.__db_create_idx_stmnt__)
            self.__db_con__.exec(self.__db_create_view_stmnt_qso_count__)

            if self.__db_con__.lastError().text():
                raise DatabaseOpenException(self.__db_con__.lastError().text())

            if not self.checkDB(db_file):
                return

            model = TranslatedTableModel(self, self.__db_con__,
                                         status_cols=tuple(range(self.__sql_cols__.index('qsl_sent'),
                                                                 self.__sql_cols__.index('hamqth') + 1)),
                                         cols=self.__sql_cols__,
                                         prop_tr=self.prop)
            model.setTable('qsos')
            self.QSOTableView.setModel(model)

            for c in self.__sql_cols__:
                model.setHeaderData(self.__sql_cols__.index(c),
                                    QtCore.Qt.Orientation.Horizontal,
                                    self.__header_map__[c])

            self.refreshTableView()

            self.log.info(f'Opened database {db_file}')
            self.settings.setValue('lastDatabase', db_file)
            self.setWindowTitle(self.programName + ' - ' + db_file)

            self.actionExport.setEnabled(True)
            self.actionExport_TB.setEnabled(True)
            self.actionChange_log_entry.setEnabled(True)
            self.actionDelete_log_entry.setEnabled(True)
            self.actionCopy.setEnabled(True)
            self.actionUpload_logs_to_LoTW.setEnabled(True)
            self.actionCheck_LoTW_QSL.setEnabled(True)
            self.actionUpload_to_eQSL.setEnabled(True)
            self.actionCheck_eQSL.setEnabled(True)
            self.actionDownload_eQSL.setEnabled(True)
            self.actionUpload_to_HamQTH.setEnabled(True)
            self.actionShow_statistics.setEnabled(True)
            self.actionShow_statistics_TB.setEnabled(True)
            if OPTION_QRCODEREADER:
                self.actionRead_DARC_QSL_QR_Code.setEnabled(True)
                self.actionRead_DARC_QSL_QR_Code_TB.setEnabled(True)
        except DatabaseOpenException as exc:
            self.log.exception(exc)
            if db_file == self.settings.value('lastDatabase', None):
                self.settings.setValue('lastDatabase', None)

            QtWidgets.QMessageBox.critical(
                self,
                f'{self.programName} - {self.tr("Error")}',
                str(exc))

    def refreshTableView(self, sort: bool = True):
        self.setFilter()

        hidden_cols = self.settings.value('ui/hidden_cols', '').split(',')
        for i in range(len(self.__headers__)):
            if str(i + 1) in hidden_cols:
                self.QSOTableView.hideColumn(i)
            else:
                self.QSOTableView.showColumn(i)

        if sort and self.settings.value('ui/sort_col', self.tr('Date/Time start')) in self.__headers__:
            sort_order = QtCore.Qt.SortOrder.AscendingOrder
            if self.settings.value('ui/sort_order', 'ASC') == 'DSC':
                sort_order = QtCore.Qt.SortOrder.DescendingOrder

            self.QSOTableView.sortByColumn(self.__headers__.index(self.settings.value('ui/sort_col',
                                                                                      self.tr('Date/Time start'))),
                                           sort_order)

        self.QSOTableView.resizeColumnsToContents()

    def getQueryStr(self):
        view_filter = self.getFilter()
        if view_filter:
            return f"SELECT * FROM qsos WHERE {view_filter}"
        else:
            return "SELECT * FROM qsos"

    def getFilter(self):
        if self.__table_filter__:
            return self.__table_filter__
        else:
            recent_filter = self.settings.value('ui/recent_qsos', self.tr('Show all'))
            if recent_filter == self.tr('Last week'):
                self.__recent_filter__ = "DATE(date_time) > DATE('now', '-7 days')"
            elif recent_filter == self.tr('Last month'):
                self.__recent_filter__ = "DATE(date_time) > DATE('now', '-31 days')"
            elif recent_filter == self.tr('Last half year'):
                self.__recent_filter__ = "DATE(date_time) > DATE('now', '-183 days')"
            elif recent_filter == self.tr('Last year'):
                self.__recent_filter__ = "DATE(date_time) > DATE('now', '-365 days')"
            else:
                self.__recent_filter__ = ""
            return self.__recent_filter__

    def setFilter(self):
        if self.QSOTableView.model():
            self.QSOTableView.model().setFilter(self.getFilter())
            self.refreshFilterStatus()

    def refreshFilterStatus(self):
        qsos = 0
        filtered_qsos = 0
        if self.__db_con__:
            query = self.__db_con__.exec('SELECT * FROM qso_count')
            if query.next():
                qsos = query.value(0)

            if self.__table_filter__ or self.__recent_filter__:
                filter_str = self.__table_filter__ if self.__table_filter__ else self.__recent_filter__
                query = self.__db_con__.exec('SELECT COUNT(id) FROM qsos WHERE ' + filter_str)
                if query.next():
                    filtered_qsos = query.value(0)
            else:
                filtered_qsos = qsos

        self.filterDockWidget.setWindowTitle(f'{self.tr("Filter")}: {filtered_qsos} / {qsos} QSOs')

    def setTableFilter(self):
        filter_set = []

        if self.fDateFromLineEdit.text():
            if check_format(REGEX_DATE, self.fDateFromLineEdit.text()):
                filter_set.append(f"DATE(date_time) >= DATE('{self.fDateFromLineEdit.text()}')")
                self.fDateFromLineEdit.setPalette(ColorPalettes.PaletteDefault)
            else:
                self.fDateFromLineEdit.setPalette(ColorPalettes.PaletteFaulty)
        else:
            self.fDateFromLineEdit.setPalette(ColorPalettes.PaletteDefault)

        if self.fDateToLineEdit.text():
            if check_format(REGEX_DATE, self.fDateToLineEdit.text()):
                filter_set.append(f"DATE(date_time) <= DATE('{self.fDateToLineEdit.text()}')")
                self.fDateToLineEdit.setPalette(ColorPalettes.PaletteDefault)
            else:
                self.fDateToLineEdit.setPalette(ColorPalettes.PaletteFaulty)
        else:
            self.fDateToLineEdit.setPalette(ColorPalettes.PaletteDefault)

        if self.fCallsignLineEdit.text():
            filter_set.append(f'call_sign LIKE "%{self.fCallsignLineEdit.text()}%"')
        if self.fBandComboBox.currentText():
            filter_set.append(f'band = "{self.fBandComboBox.currentText()}"')
        if self.fModeComboBox.currentText():
            filter_set.append(f'mode = "{self.fModeComboBox.currentText()}"')

        if self.fQSLsComboBox.currentIndex() > 0:
            filter_set.append(f'qsl_sent {"!" if self.fQSLsComboBox.currentIndex() == 2 else ""}= "Y"')
        if self.fQSLrComboBox.currentIndex() > 0:
            filter_set.append(f'qsl_rcvd {"!" if self.fQSLrComboBox.currentIndex() == 2 else ""}= "Y"')
        if self.feQSLsComboBox.currentIndex() > 0:
            filter_set.append(f'eqsl_sent {"!" if self.feQSLsComboBox.currentIndex() == 2 else ""}= "Y"')
        if self.feQSLrComboBox.currentIndex() > 0:
            filter_set.append(f'eqsl_rcvd {"!" if self.feQSLrComboBox.currentIndex() == 2 else ""}= "Y"')
        if self.fLoTWsComboBox.currentIndex() > 0:
            filter_set.append(f'lotw_sent {"!" if self.fLoTWsComboBox.currentIndex() == 2 else ""}= "Y"')
        if self.fLoTWrComboBox.currentIndex() > 0:
            filter_set.append(f'lotw_rcvd {"!" if self.fLoTWrComboBox.currentIndex() == 2 else ""}= "Y"')

        if self.fContestComboBox.currentText():
            cntst_id = self.fContestComboBox.currentText()
            if cntst_id in CONTEST_IDS:
                cntst_id = CONTEST_IDS[cntst_id]
            filter_set.append(f'contest_id = "{cntst_id}"')

        self.__table_filter__ = ' AND '.join(filter_set)

        self.setFilter()

    def resetTableFilter(self):
        self.__table_filter__ = ''
        self.setFilter()

        self.fDateFromLineEdit.clear()
        self.fDateFromLineEdit.setPalette(ColorPalettes.PaletteDefault)
        self.fDateToLineEdit.clear()
        self.fDateToLineEdit.setPalette(ColorPalettes.PaletteDefault)
        self.fCallsignLineEdit.clear()
        self.fBandComboBox.setCurrentIndex(0)
        self.fModeComboBox.setCurrentIndex(0)
        self.fQSLsComboBox.setCurrentIndex(0)
        self.fQSLrComboBox.setCurrentIndex(0)
        self.feQSLsComboBox.setCurrentIndex(0)
        self.feQSLrComboBox.setCurrentIndex(0)
        self.fLoTWsComboBox.setCurrentIndex(0)
        self.fLoTWrComboBox.setCurrentIndex(0)
        self.fContestComboBox.setCurrentIndex(0)

    def ctrlHamlib(self, start):
        try:
            self.__rigctl__.ctrlRigctld(start)
        except (RigctldNotConfiguredException, CATSettingsMissingException, RigctldExecutionException):
            self.log.error(f'rigctld is not properly configured')
        except Exception as exc:
            self.log.exception(exc)

    def logQSO(self):
        if not self.__db_con__.isOpen():
            self.selectDB()
            if not self.__db_con__.isOpen():
                QtWidgets.QMessageBox.warning(self, self.tr('Log QSO'),
                                              self.tr('No database opened for logging'))
                return

        self.qso_form.clear()
        self.qso_form.setChangeMode(False)
        self.qso_form.reset()
        self.qsoDockWidget.setVisible(True)
        self.qso_form.callSignLineEdit.setFocus()

    def fetchQSO(self):
        """Fetch a QSO from QSO form"""
        if not self.__db_con__.isOpen():
            self.selectDB()
            if not self.__db_con__.isOpen():
                QtWidgets.QMessageBox.warning(self, self.tr('Saving QSO'),
                                              self.tr('No database opened for saving'))
                return

        self.log.info('Logging QSO...')
        query = QtSql.QSqlQuery(self.__db_con__)
        query.prepare(self.__db_insert_stmnt__)
        values = self.qso_form.values
        for i, val in enumerate(values):
            query.bindValue(i, val)
        query.exec()
        if query.lastError().text():
            raise Exception(query.lastError().text())

        self.__db_con__.commit()
        self.refreshTableView(sort=False)

        self.hamlib_error.setText('')  # TODO: Why???
        self.qsoAdded.emit()

        qso = {
            'CALL': values[self.__sql_cols__.index('call_sign') - 1],
            'NAME_INTL': values[self.__sql_cols__.index('name') - 1],
            'QTH_INTL': values[self.__sql_cols__.index('qth') - 1],
            'GRIDSQUARE': values[self.__sql_cols__.index('locator') - 1],
            'CONTEST_ID': values[self.__sql_cols__.index('contest_id') - 1],
            'SRX_STRING': values[self.__sql_cols__.index('crx_data') - 1],
        }
        self.addQSOToCallbook(qso)

    def selectedQSOIds(self) -> Iterator[int]:
        """Fetch the QSO ID for selected"""
        yielded_ids: list[int] = []

        for i in self.QSOTableView.selectedIndexes():
            qso_id = self.QSOTableView.model().data(i.siblingAtColumn(0))

            if qso_id in yielded_ids or qso_id is None:
                continue
            else:
                yielded_ids.append(qso_id)
                yield qso_id

    def changeContestID(self):
        """Change the contest ID for selected QSOs"""
        contest, ok = QtWidgets.QInputDialog.getItem(self, self.tr('Change contest'),
                                                     self.tr('Select new contest'),
                                                     list(CONTEST_NAMES.values()),
                                                     editable=False)
        if not ok:
            return

        for q in self.selectedQSOIds():
            self.updateQSOField('contest_id', q, CONTEST_IDS[contest])

        self.refreshTableView()

    def selectedQSOs(self) -> Iterator[dict[str, str]]:
        """Fetch the QSO data for selected QSOs"""
        for qso_id in self.selectedQSOIds():
            query = self.__db_con__.exec(f'SELECT * FROM qsos WHERE id = {qso_id}')
            if query.lastError().text():
                raise Exception(query.lastError().text())

            values = []
            while query.next():
                for col in range(len(self.__sql_cols__)):
                    values.append(query.value(col))
                break
            yield dict(zip(self.__sql_cols__, values))

    def copyToClipboard(self):
        table = []
        for i, qso in enumerate(self.selectedQSOs()):
            if i == 0:
                table.append('\t'.join([str(c) for c in qso.keys()]))
            table.append('\t'.join([str(c) for c in qso.values()]))

        clip = QtWidgets.QApplication.clipboard()
        clip.setText('\r\n'.join(table))

    def deleteQSO(self):
        if self.QSOTableView.selectedIndexes():
            res = QtWidgets.QMessageBox.question(self, self.tr('Delete QSO'),
                                                 self.tr('Do you really want to delete the selected QSO(s)?'),
                                                 defaultButton=QtWidgets.QMessageBox.StandardButton.No)

            if res == QtWidgets.QMessageBox.StandardButton.Yes:
                for qso_id in self.selectedQSOIds():
                    self.log.info(f'Deleting QSO #{qso_id}...')
                    query = QtSql.QSqlQuery(self.__db_con__)
                    query.prepare('DELETE FROM qsos where id == ?')
                    query.bindValue(0, qso_id)
                    query.exec()

                    if query.lastError().text():
                        raise Exception(query.lastError().text())

                self.__db_con__.commit()
                self.refreshTableView(sort=False)
                self.qsoDeleted.emit()

    def changeQSO(self):
        if self.QSOTableView.selectedIndexes():
            self.qso_form.clear()
            self.qsoDockWidget.setVisible(True)
            self.qso_form.setChangeMode()

            i = self.QSOTableView.selectedIndexes().pop(0)
            qso_id = self.QSOTableView.model().data(i.siblingAtColumn(0))

            query = self.__db_con__.exec(f'SELECT * FROM qsos WHERE id = {qso_id}')
            if query.lastError().text():
                raise Exception(query.lastError().text())

            values = []
            while query.next():
                for col in range(len(self.__sql_cols__)):
                    values.append(query.value(col))
                break
            self.qso_form.values = dict(zip(self.__sql_cols__, values))

    def updateQSO(self, qso_id: int):
        self.log.info(f'Changing QSO {qso_id}...')
        values = self.qso_form.values
        values += (qso_id,)

        query = QtSql.QSqlQuery(self.__db_con__)
        query.prepare(self.__db_update_stmnt__)

        for col, val in enumerate(values):
            query.bindValue(col, val)
        query.exec()
        if query.lastError().text():
            raise Exception(query.lastError().text())

        self.__db_con__.commit()
        self.refreshTableView(sort=False)
        self.qso_form.setChangeMode(False)
        self.qsoChanged.emit()

        qso = {
            'CALL': values[self.__sql_cols__.index('call_sign') - 1],
            'NAME_INTL': values[self.__sql_cols__.index('name') - 1],
            'QTH_INTL': values[self.__sql_cols__.index('qth') - 1],
            'GRIDSQUARE': values[self.__sql_cols__.index('locator') - 1],
        }

        qso_date = values[self.__sql_cols__.index('date_time') - 1].split(' ')[0]
        datetime.date.today().strftime('%Y-%m-%d')
        if qso_date == datetime.date.today().strftime('%Y-%m-%d'):
            qso['CONTEST_ID'] = values[self.__sql_cols__.index('contest_id') - 1]
            qso['SRX_STRING'] = values[self.__sql_cols__.index('crx_data') - 1]

        self.addQSOToCallbook(qso)

    def clearQSOForm(self):
        self.qso_form.clear()
        self.qso_form.setChangeMode(False)

    def export(self):
        qso_filters = [
            self.tr('All QSOs'),
            self.tr('Filtered QSOs'),  # Must be second element for pre selection
            self.tr('Selected QSOs'),
        ]
        qso_filter, ok = QtWidgets.QInputDialog.getItem(self, self.tr('QSO export'),
                                                        self.tr('Select filter'),
                                                        qso_filters,
                                                        # Pre select second element "Filetered QSOs" if option is set
                                                        int(self.settings.value('imp_exp/only_recent', 0)))
        if not ok:
            return

        query_str = self.__db_select_stmnt__
        if qso_filters.index(qso_filter) == 1:
            query_str = self.getQueryStr()
        elif qso_filters.index(qso_filter) == 2:
            sel_qsos = [str(qso) for qso in self.selectedQSOIds()]
            query_str = f"SELECT * FROM qsos WHERE id IN ({','.join(sel_qsos)})"

        exp_formats = {
            self.tr('ADIF 3 (*.adi *.adif)'): self.exportADIF,
            self.tr('ADIF 3 zipped (*.zip)'): self.exportADIF,
            self.tr('ADIF 3 XML (*.adx)'): self.exportADIF,
            self.tr('CSV-File (*.csv)'): self.exportCSV,
        }

        if OPTION_OPENPYXL:
            exp_formats[self.tr('Excel-File (*.xlsx)')] = self.exportExcel

        res = QtWidgets.QFileDialog.getSaveFileName(
            self,
            self.tr('Select export file'),
            os.path.join(self.settings.value('lastExportDir', os.path.abspath(os.curdir)),
                         datetime.date.today().strftime('%Y-%m-%d ') + qso_filter),
            ';;'.join(exp_formats.keys()))

        if res[0]:
            try:
                # noinspection PyArgumentList
                exp_formats[res[1]](res[0], query_str)
            except Exception as exc:
                self.log.exception(exc)

            self.settings.setValue('lastExportDir', os.path.abspath(os.path.dirname(res[0])))

    def exportCSV(self, file: str, query_str: str):
        self.log.info('Exporting to CSV...')

        try:
            with open(file, 'w', newline='', encoding='utf-8') as cf:
                writer = csv.writer(cf, delimiter=self.settings_form.csv_delimiter)

                # Write header
                writer.writerow(self.__headers__)

                # Write content
                query = self.__db_con__.exec(query_str)

                if query.lastError().text():
                    raise Exception(query.lastError().text())

                while query.next():
                    row = []
                    for i in range(len(self.__sql_cols__)):
                        row.append(query.value(i))
                    writer.writerow(row)

            self.log.info(f'Saved "{file}"')
        except OSError as e:
            self.log.critical(e)
            QtWidgets.QMessageBox.critical(
                self,
                f'{self.programName} - {self.tr("Error")}',
                str(e))

    def exportExcel(self, file: str, query_str: str):
        self.log.info('Exporting to XLSX...')
        xl_wb = openpyxl.Workbook()
        xl_wb.properties.title = self.tr('Exported QSO log')
        xl_wb.properties.description = f'{self.programName} {self.programVersion}'
        xl_wb.properties.creator = os.getlogin()

        xl_ws = xl_wb.active
        xl_ws.title = 'QSOs'
        xl_ws.freeze_panes = 'A2'

        col_widths = []

        # Write header
        column = 1
        for h in self.__headers__:
            cell = xl_ws.cell(column=column, row=1, value=h)
            col_widths.append(len(h))  # Initialise width approximation
            cell.font = Font(bold=True)
            column += 1

        # Write content
        query = self.__db_con__.exec(query_str)
        row = 2
        while query.next():
            for col in range(len(self.__headers__)):
                val = query.value(col)
                xl_ws.cell(column=col + 1, row=row, value=val)
                val_len = len(str(val))
                if col_widths[col] < val_len:
                    col_widths[col] = val_len
            row += 1

        # Set auto filter
        # noinspection PyUnresolvedReferences
        xl_ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(self.__headers__))}1'

        # Fit size to content width approximation
        for c, w in zip(range(1, len(col_widths) + 1), col_widths):
            # Add 5 due to Excel filter drop down
            # noinspection PyUnresolvedReferences
            xl_ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w + 5

        # Finally save
        try:
            xl_wb.save(file)
            self.log.info(f'Saved "{file}"')
        except OSError as e:
            self.log.critical(e)
            QtWidgets.QMessageBox.critical(
                self,
                f'{self.programName} - {self.tr("Error")}',
                str(e))

    def replaceNonASCII(self, text: str) -> str:
        non_ascii = find_non_ascii(text)

        for na in non_ascii:
            if na:
                if na in self.ascii_replace:
                    text = text.replace(na, self.ascii_replace[na])
                else:
                    text = text.replace(na, '_')

        return text

    def exportADIF(self, file: str, query_str: str):
        self.log.info('Exporting to ADIF...')

        is_adx: bool = os.path.splitext(file)[-1] == '.adx'
        doc = self.build_adif_export(query_str, is_adx)
        try:
            if is_adx:
                errors = adx.dump(file, doc, False)
                for err in errors:
                    self.log.warning('ADX validation error: ' + str(err))
                if errors:
                    QtWidgets.QMessageBox.warning(
                        self,
                        f'{self.programName} - {self.tr("ADX-Export")}',
                        self.tr('ADX validation detected one or more error(s)\nSee log for detail'))
            elif os.path.splitext(file)[-1] == '.zip':
                with zipfile.ZipFile(file, 'w', zipfile.ZIP_DEFLATED) as z_file:
                    a_zinfo = zipfile.ZipInfo(os.path.splitext(os.path.basename(file))[0] + '.adi',
                                              datetime.datetime.now().timetuple()[:6])
                    a_zinfo.compress_type = zipfile.ZIP_DEFLATED
                    with z_file.open(a_zinfo, 'w') as a_file:
                        a_file.write(adi.dumps(doc, 'ADIF Export by DragonLog').encode())
            else:
                adi.dump(file, doc, 'ADIF Export by DragonLog')

            self.log.info(f'Saved "{file}"')
        except Exception as exc:
            self.log.exception(exc)
            QtWidgets.QMessageBox.critical(
                self,
                f'{self.programName} - {self.tr("Error")}',
                str(exc))

    def build_adif_export(self, query_str, is_adx=False, include_id=False) -> dict[
        str, dict[str, str] | list[dict[str, str]] | None]:
        """Build a dict for ADIF export from database query"""

        doc: dict[
            str, dict[str, str] | list[dict[str, str]] | None] = {
            'HEADER':
                {
                    'ADIF_VER': '3.1.4',
                    'PROGRAMID': self.programName,
                    'PROGRAMVERSION': self.programVersion,
                    'CREATED_TIMESTAMP': QtCore.QDateTime.currentDateTimeUtc().toString('yyyyMMdd HHmmss')
                },
            'RECORDS': None,
        }
        records: list[dict[str, str]] = []

        query = self.__db_con__.exec(query_str)
        if query.lastError().text():
            raise Exception(query.lastError().text())
        while query.next():
            band = query.value(self.__sql_cols__.index('band'))

            if band == '11m' and not self.settings.value('station_cb/cb_exp_adif', 0):
                continue

            qso_date, qso_time = query.value(self.__sql_cols__.index('date_time')).split()

            record = {'QSO_DATE': qso_date.replace('-', ''),
                      'TIME_ON': qso_time.replace(':', ''),
                      'APP': []}

            if include_id and query.value(self.__sql_cols__.index('id')):
                if is_adx:
                    record['APP'].append({'@PROGRAMID': 'DRAGONLOG',
                                          '@FIELDNAME': 'QSOID',
                                          '@TYPE': 'N',
                                          '$': query.value(self.__sql_cols__.index('id'))})
                else:
                    record['APP_DRAGONLOG_QSOID'] = query.value(self.__sql_cols__.index('id'))
            if query.value(self.__sql_cols__.index('date_time_off')):
                qso_date_off, qso_time_off = query.value(self.__sql_cols__.index('date_time_off')).split()
                record['QSO_DATE_OFF'] = qso_date_off.replace('-', '')
                record['TIME_OFF'] = qso_time_off.replace(':', '')

            if query.value(self.__sql_cols__.index('call_sign')):
                if band == '11m' and is_adx:
                    record['APP'].append({'@PROGRAMID': 'DRAGONLOG',
                                          '@FIELDNAME': 'CBCALL',
                                          '@TYPE': 'I',
                                          '$': query.value(self.__sql_cols__.index('call_sign'))})
                else:
                    record['CALL'] = self.replaceNonASCII(query.value(self.__sql_cols__.index('call_sign')))
            if query.value(self.__sql_cols__.index('name')):
                record['NAME'] = self.replaceNonASCII(query.value(self.__sql_cols__.index('name')))
                record['NAME_INTL'] = query.value(self.__sql_cols__.index('name'))
            if query.value(self.__sql_cols__.index('qth')):
                record['QTH'] = self.replaceNonASCII(query.value(self.__sql_cols__.index('qth')))
                record['QTH_INTL'] = query.value(self.__sql_cols__.index('qth'))
            if query.value(self.__sql_cols__.index('locator')):
                record['GRIDSQUARE'] = query.value(self.__sql_cols__.index('locator'))
            if query.value(self.__sql_cols__.index('rst_sent')):
                record['RST_SENT'] = query.value(self.__sql_cols__.index('rst_sent'))
            if query.value(self.__sql_cols__.index('rst_rcvd')):
                record['RST_RCVD'] = query.value(self.__sql_cols__.index('rst_rcvd'))
            if band:
                if band == '11m':
                    if is_adx:
                        record['APP'].append({'@PROGRAMID': 'DRAGONLOG',
                                              '@FIELDNAME': 'CBQSO',
                                              '@TYPE': 'B',
                                              '$': 'Y'})
                        if query.value(self.__sql_cols__.index("channel")):
                            record['APP'].append({'@PROGRAMID': 'DRAGONLOG',
                                                  '@FIELDNAME': 'CBCHANNEL',
                                                  '@TYPE': 'N',
                                                  '$': str(query.value(self.__sql_cols__.index("channel")))})
                    else:
                        record['BAND'] = band.upper()
                        record['APP_DRAGONLOG_CBQSO'] = 'Y'
                        if query.value(self.__sql_cols__.index("channel")):
                            record['APP_DRAGONLOG_CBCHANNEL'] = query.value(self.__sql_cols__.index("channel"))
                else:
                    record['BAND'] = band.upper()
            if query.value(self.__sql_cols__.index('mode')):
                record['MODE'] = query.value(self.__sql_cols__.index('mode'))
            if query.value(self.__sql_cols__.index("freq")):
                record['FREQ'] = f'{query.value(self.__sql_cols__.index("freq")) / 1000:0.6f}'.rstrip('0').rstrip('.')
            if query.value(self.__sql_cols__.index('power')):
                record['TX_PWR'] = query.value(self.__sql_cols__.index('power'))
            if query.value(self.__sql_cols__.index('propagation')):
                record['PROP_MODE'] = query.value(self.__sql_cols__.index('propagation'))
            if query.value(self.__sql_cols__.index('own_callsign')):
                if band == '11m' and is_adx:
                    record['APP'].append({'@PROGRAMID': 'DRAGONLOG',
                                          '@FIELDNAME': 'CBOWNCALL',
                                          '@TYPE': 'I',
                                          '$': query.value(self.__sql_cols__.index('own_callsign'))})
                else:
                    record['STATION_CALLSIGN'] = self.replaceNonASCII(
                        query.value(self.__sql_cols__.index('own_callsign')))
            if query.value(self.__sql_cols__.index('own_name')):
                record['MY_NAME'] = self.replaceNonASCII(query.value(self.__sql_cols__.index('own_name')))
                record['MY_NAME_INTL'] = query.value(self.__sql_cols__.index('own_name'))
            if query.value(self.__sql_cols__.index('own_qth')):
                record['MY_CITY'] = self.replaceNonASCII(query.value(self.__sql_cols__.index('own_qth')))
                record['MY_CITY_INTL'] = query.value(self.__sql_cols__.index('own_qth'))
            if query.value(self.__sql_cols__.index('own_locator')):
                record['MY_GRIDSQUARE'] = query.value(self.__sql_cols__.index('own_locator'))
            if query.value(self.__sql_cols__.index('radio')):
                record['MY_RIG'] = self.replaceNonASCII(query.value(self.__sql_cols__.index('radio')))
                record['MY_RIG_INTL'] = query.value(self.__sql_cols__.index('radio'))
            if query.value(self.__sql_cols__.index('antenna')):
                record['MY_ANTENNA'] = self.replaceNonASCII(query.value(self.__sql_cols__.index('antenna')))
                record['MY_ANTENNA_INTL'] = query.value(self.__sql_cols__.index('antenna'))
            if query.value(self.__sql_cols__.index('dist')):
                record['DISTANCE'] = query.value(self.__sql_cols__.index('dist'))
            if query.value(self.__sql_cols__.index('remarks')) and bool(
                    self.settings.value('imp_exp/own_notes_adif', 0)):
                remarks = query.value(self.__sql_cols__.index('remarks'))
                if platform.system() == 'Linux':
                    remarks = remarks.replace('\n', '\r\n')
                record['NOTES'] = self.replaceNonASCII(remarks)
                record['NOTES_INTL'] = remarks
            if query.value(self.__sql_cols__.index('comments')):
                comment = query.value(self.__sql_cols__.index('comments'))
                record['COMMENT'] = self.replaceNonASCII(comment)
                record['COMMENT_INTL'] = comment
            if query.value(self.__sql_cols__.index('qsl_via')):
                record['QSL_VIA'] = self.replaceNonASCII(query.value(self.__sql_cols__.index('qsl_via')))
            if query.value(self.__sql_cols__.index('qsl_path')):
                record['QSL_SENT_VIA'] = query.value(self.__sql_cols__.index('qsl_path'))
            if query.value(self.__sql_cols__.index('qsl_msg')):
                qsl_msg = query.value(self.__sql_cols__.index('qsl_msg'))
                if platform.system() == 'Linux':
                    qsl_msg = qsl_msg.replace('\n', '\r\n')
                record['QSLMSG'] = self.replaceNonASCII(qsl_msg)
                record['QSLMSG_INTL'] = qsl_msg
            if query.value(self.__sql_cols__.index('qsl_sent')):
                record['QSL_SENT'] = query.value(self.__sql_cols__.index('qsl_sent'))
            if query.value(self.__sql_cols__.index('qsl_rcvd')):
                record['QSL_RCVD'] = query.value(self.__sql_cols__.index('qsl_rcvd'))
            if query.value(self.__sql_cols__.index('eqsl_sent')):
                record['EQSL_QSL_SENT'] = query.value(self.__sql_cols__.index('eqsl_sent'))
            if query.value(self.__sql_cols__.index('eqsl_rcvd')):
                record['EQSL_QSL_RCVD'] = query.value(self.__sql_cols__.index('eqsl_rcvd'))
            if query.value(self.__sql_cols__.index('lotw_sent')):
                record['LOTW_QSL_SENT'] = query.value(self.__sql_cols__.index('lotw_sent'))
            if query.value(self.__sql_cols__.index('lotw_rcvd')):
                record['LOTW_QSL_RCVD'] = query.value(self.__sql_cols__.index('lotw_rcvd'))
            if query.value(self.__sql_cols__.index('hamqth')):
                record['HAMQTH_QSO_UPLOAD_STATUS'] = query.value(self.__sql_cols__.index('hamqth'))
            if query.value(self.__sql_cols__.index('contest_id')):
                c_id = query.value(self.__sql_cols__.index('contest_id'))
                # Try to get different ADIF ID or use internal
                record['CONTEST_ID'] = CONTEST_DL_ADIF.get(c_id, c_id)
            if query.value(self.__sql_cols__.index('ctx_qso_id')):
                record['STX'] = query.value(self.__sql_cols__.index('ctx_qso_id'))
            if query.value(self.__sql_cols__.index('crx_qso_id')):
                record['SRX'] = query.value(self.__sql_cols__.index('crx_qso_id'))
            if query.value(self.__sql_cols__.index('crx_data')):
                record['SRX_STRING'] = query.value(self.__sql_cols__.index('crx_data'))
            if query.value(self.__sql_cols__.index('event')):
                event = query.value(self.__sql_cols__.index('event'))
                event_ascii = self.replaceNonASCII(event)
                record['MY_SIG_INTL'] = event
                record['MY_SIG'] = event_ascii
                record['SIG_INTL'] = event
                record['SIG'] = event_ascii
            if query.value(self.__sql_cols__.index('evt_tx_exch')):
                exch = query.value(self.__sql_cols__.index('evt_tx_exch'))
                record['MY_SIG_INFO_INTL'] = exch
                record['MY_SIG_INFO'] = self.replaceNonASCII(exch)
            if query.value(self.__sql_cols__.index('evt_rx_exch')):
                exch = query.value(self.__sql_cols__.index('evt_rx_exch'))
                record['SIG_INFO_INTL'] = exch
                record['SIG_INFO'] = self.replaceNonASCII(exch)

            if not record['APP']:
                record.pop('APP')

            records.append(record)
        doc['RECORDS'] = records
        return doc

    def eqslUpload(self):
        for qso_id in self.selectedQSOIds():
            adif_doc = self.build_adif_export(f"SELECT * FROM qsos WHERE id = {qso_id} AND band != '11m'")
            if not adif_doc['RECORDS']:
                self.log.info(f'Skipped CB QSO #{qso_id}')
                continue

            rec = adif_doc['RECORDS'][0]

            if rec.get('EQSL_QSL_SENT', 'N') != 'N':
                self.log.info(f'QSO #{qso_id} already uploaded to eQSL')
                continue

            eqsl_sent = 'N'
            try:
                self.eqsl.upload_log(self.settings.value('eqsl/username', ''),
                                     self.settings_form.eqslPassword(),
                                     rec)
                eqsl_sent = 'Y'
                self.log.info(f'Uploaded QSO #{qso_id} to eQSL')
            except EQSLLoginException:
                QtWidgets.QMessageBox.warning(self, self.tr('Upload eQSL error'),
                                              self.tr('Login failed for user') + ': ' + self.settings.value(
                                                  'eqsl/username', ''))
                break
            except EQSLADIFFieldException as exc:
                self.log.warning(f'A field is missing in QSO #{qso_id} for eQSL check: "{exc.args[0]}"')
            except EQSLQSODuplicateException:
                eqsl_sent = 'Y'
                self.log.info(f'QSO #{qso_id} is a dublicate, already uploaded to eQSL')
            except EQSLUserCallMatchException:
                QtWidgets.QMessageBox.warning(self, self.tr('Upload eQSL error'),
                                              self.tr('User call does not match') + ': ' + self.settings.value(
                                                  'eqsl/username', ''))
                break
            except EQSLRequestException as exc:
                QtWidgets.QMessageBox.information(self, self.tr('Upload eQSL error'),
                                                  self.tr('Error on upload') + f':\n"{exc.args[0]}"')
                break
            except EQSLCommunicationException as exc:
                self.log.warning(str(exc))
                QtWidgets.QMessageBox.warning(self, self.tr('Upload eQSL error'),
                                              self.tr('eQSL request error'))
                break
            finally:
                self.updateQSOField('eqsl_sent', qso_id, eqsl_sent)
        self.refreshTableView(False)

    def eqslCheckInboxSelected(self):
        for qso_id in self.selectedQSOIds():
            res = self.eqslCheckInbox(qso_id)
            if not res:
                break
        self.refreshTableView(False)

    def eqslCheckInbox(self, qso_id) -> bool:
        """Checks eQSL Inbox
        :param qso_id: the QSO Id to check the eQSL for
        :return: True if request was a success, False if request failed"""
        adif_doc = self.build_adif_export(f"SELECT * FROM qsos WHERE id = {qso_id} AND band != '11m'")
        if not adif_doc['RECORDS']:
            self.log.info(f'Skipped CB QSO #{qso_id}')
            return False

        rec = adif_doc['RECORDS'][0]
        try:
            if int(self.settings.value('eqsl/use_qso_call', 0)) and rec.get('STATION_CALLSIGN', ''):
                username = rec['STATION_CALLSIGN']
            else:
                username = self.settings.value('eqsl/username', '')
            self.log.info(f'Checking eQSL QSO #{qso_id} for call {username}...')
            res = self.eqsl.check_inbox(username,
                                        self.settings_form.eqslPassword(),
                                        rec)
            if res:
                qso_uuid = rec['QSO_DATE'] + rec['TIME_ON'] + rec['CALL']
                self.eqsl_urls[qso_uuid] = res
                self.log.debug(f'eQSL available at "{res}"')
                self.updateQSOField('eqsl_rcvd', qso_id, 'Y')
                return True
        except EQSLLoginException as exc:
            QtWidgets.QMessageBox.warning(self, self.tr('Check eQSL Inbox error'),
                                          self.tr('Login failed for user') + ': ' + self.settings.value(
                                              'eqsl/username', '') + f'\n{exc}')
        except EQSLUserCallMatchException:
            QtWidgets.QMessageBox.warning(self, self.tr('Check eQSL Inbox error'),
                                          self.tr('User call does not match') + ': ' + self.settings.value(
                                              'eqsl/username', ''))
        except EQSLRequestException:
            self.log.info('No eQSL available')
            return True
        except EQSLADIFFieldException as exc:
            self.log.warning(f'A field is missing in QSO #{qso_id} for eQSL check: "{exc.args[0]}"')
        except EQSLCommunicationException as exc:
            self.log.warning(str(exc))
            QtWidgets.QMessageBox.warning(self, self.tr('Check eQSL Inbox error'),
                                          self.tr('eQSL request error'))
        return False

    def eqslDownload(self):
        res = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            self.tr('Select eQSL folder'),
            self.settings.value('eqsl/lastExportDir', os.path.abspath(os.curdir)))

        if not res:
            return

        for qso_id in self.selectedQSOIds():
            adif_doc = self.build_adif_export(f"SELECT * FROM qsos WHERE id = {qso_id} AND band != '11m'")
            if not adif_doc['RECORDS']:
                self.log.info(f'Skipped CB QSO #{qso_id}')
                continue

            self.log.info(f'Downloading eQSL for QSO #{qso_id}...')

            rec = adif_doc['RECORDS'][0]
            qso_uuid = rec['QSO_DATE'] + rec['TIME_ON'] + rec['CALL']

            if qso_uuid not in self.eqsl_urls:
                if not self.eqslCheckInbox(qso_id):
                    break

            if self.eqsl_urls.get(qso_uuid):
                image_type = self.eqsl_urls[qso_uuid].split('/')[-1].split('.')[-1]
                call_sign = rec['CALL']
                for c in string.punctuation:
                    call_sign = call_sign.replace(c, '_')

                image_name = (f'{rec["QSO_DATE"]} {call_sign} {rec["MODE"]} '
                              f'{rec["BAND"]}.{image_type}')
                image_path = os.path.join(res, image_name)
                self.log.debug(f'eQSL path: "{image_path}"')

                try:
                    eqsl_image = self.eqsl.receive_qsl_card(self.eqsl_urls[qso_uuid])
                    with open(image_path, 'wb') as eqslf:
                        eqslf.write(eqsl_image)
                    self.log.info(f'Stored eQSL to "{image_path}"')
                    self.settings.setValue('eqsl/lastExportDir', res)
                except Exception as exc:
                    self.log.exception(exc)
        self.refreshTableView(False)

    def lotwUpload(self):
        self.log.info('Signing and uploading to LoTW...')
        if platform.system() == 'Windows':
            stationf_path = os.path.expanduser('~/AppData/Roaming/TrustedQSL/station_data')
        elif platform.system() == 'Linux':
            stationf_path = os.path.expanduser('~/.tqsl/station_data')
        else:
            self.log.warning(f'System "{platform.system()}" is currently not supported to upload to LoTW')
            return

        if not os.path.isfile(stationf_path):
            QtWidgets.QMessageBox.warning(self, self.tr('LoTW ADIF upload'),
                                          self.tr('Missing station configuration in TQSL'))
            return

        try:
            with open(stationf_path) as st_f:
                station_xml = st_f.read()

            station_def = xmltodict.parse(station_xml)

            stations = {}
            for st in station_def['StationDataFile']['StationData']:
                if st['CALL'] == '[None]':
                    continue
                stations[f"{st['CALL']} - {st['@name']}"] = {
                    'name': st['@name'],
                    'locator': st['GRIDSQUARE'],
                    'call': st['CALL'],
                }
        except Exception as exc:
            self.log.exception(exc)
            return

        station, ok = QtWidgets.QInputDialog.getItem(self, self.tr('LoTW ADIF upload'),
                                                     self.tr('Select station'),
                                                     sorted(stations))
        if not ok:
            return

        self.log.info(f'Selected station "{station}"')
        locator = stations[station]['locator'].upper()
        call = stations[station]['call'].upper()
        doc = self.build_adif_export(f"SELECT * FROM qsos "
                                     f"WHERE band != '11m' AND upper(own_locator) LIKE '{locator}%'"
                                     f"AND own_callsign = '{call}'"
                                     f"AND (lotw_sent != 'Y' OR lotw_sent is NULL)",
                                     include_id=True)
        if len(doc['RECORDS']) < 1:
            QtWidgets.QMessageBox.warning(self, self.tr('LoTW ADIF upload'),
                                          self.tr('No records for location') + f': "{locator}"')
            return

        passwd, ok = '', False
        if self.settings.value('lotw/cert_needs_pwd', 0):
            passwd, ok = QtWidgets.QInputDialog.getText(self, self.tr('LoTW ADIF upload'),
                                                        self.tr('TQSL signature password'),
                                                        echo=QtWidgets.QLineEdit.EchoMode.PasswordEchoOnEdit)
            if not ok:
                return

        try:
            if self.lotw.upload_log(stations[station]['name'], doc, passwd if ok else ''):
                for rec in doc['RECORDS']:
                    update_stmnt = "UPDATE qsos SET lotw_sent='Y', lotw_rcvd='R' WHERE id = ?"

                    self.log.debug(f"Changing LoTW status for QSO {rec['APP_DRAGONLOG_QSOID']}...")
                    query = QtSql.QSqlQuery(self.__db_con__)
                    query.prepare(update_stmnt)
                    query.bindValue(0, rec['APP_DRAGONLOG_QSOID'])
                    query.exec()
                    if query.lastError().text():
                        raise Exception(query.lastError().text())

                self.__db_con__.commit()
                self.refreshTableView()
        except LoTWADIFFieldException as exc:
            QtWidgets.QMessageBox.critical(self, self.tr('LoTW ADIF upload'),
                                           self.tr('A field is missing for upload') + f':\n"{exc.args[0]}"')
        except LoTWCommunicationException:
            QtWidgets.QMessageBox.critical(self, self.tr('LoTW ADIF upload'),
                                           self.tr('Connection error or network unreachable'))
        except LoTWRequestException:
            QtWidgets.QMessageBox.critical(self, self.tr('LoTW ADIF upload'),
                                           self.tr('Server rejected log'))

    def lotwCheckQSL(self):
        lotw = LoTW(self.log)

        for qso_id in self.selectedQSOIds():
            adif_doc = self.build_adif_export(f"SELECT * FROM qsos WHERE id = {qso_id} AND band != '11m'")
            if not adif_doc['RECORDS']:
                self.log.info(f'Skipped CB QSO #{qso_id}')
                continue

            rec = adif_doc['RECORDS'][0]
            lotw_sent = 'Y' if rec.get('LOTW_QSL_SENT', 'N') in ('Y', 'R') else 'N'
            lotw_rcvd = 'N'

            self.log.info(f'Checking LoTW QSL for #{qso_id}...')
            try:
                rcvd = lotw.check_inbox(self.settings.value('lotw/username', ''),
                                        self.settings_form.lotwPassword(),
                                        rec)
                lotw_rcvd = 'Y' if rcvd else 'N'
                lotw_sent = 'Y'

                self.log.info(f'LoTW: QSL for QSO #{qso_id}' if rcvd else f'LoTW: No QSL for QSO #{qso_id}')
            except LoTWNoRecordException:
                self.log.info(f'LoTW: No record available for QSO #{qso_id}')
            except LoTWCommunicationException:
                QtWidgets.QMessageBox.warning(self, self.tr('Check LoTW Inbox error'),
                                              self.tr('Server communication error'))
                break
            except LoTWRequestException as exc:
                QtWidgets.QMessageBox.warning(self, self.tr('Check LoTW Inbox error'),
                                              self.tr('Bad request result') + f'\n{exc}')
                break
            except LoTWLoginException as exc:
                QtWidgets.QMessageBox.warning(self, self.tr('Check LoTW Inbox error'),
                                              self.tr('Login failed for user') + ': ' + self.settings.value(
                                                  'lotw/username', '') + f'\n{exc}')
                break
            finally:
                self.updateQSOField('lotw_sent', qso_id, lotw_sent)
                self.updateQSOField('lotw_rcvd', qso_id, lotw_rcvd)
        self.refreshTableView(False)

    def uploadToHamQTH(self):
        logbook = None

        for qso_id in self.selectedQSOIds():
            if not logbook:
                logbook = HamQTHCallBook(self.log,
                                         f'{self.programName}-{self.programVersion}',
                                         )

            adif_doc = self.build_adif_export(f"SELECT * FROM qsos WHERE id = {qso_id} AND band != '11m'")
            if not adif_doc['RECORDS']:
                self.log.info(f'Skipped CB QSO #{qso_id}')
                continue

            if adif_doc['RECORDS'][0].get('HAMQTH_QSO_UPLOAD_STATUS', 'N') != 'N':
                self.log.debug(f'{qso_id} already uploaded to HamQTH')
                continue

            self.log.info(f'Uploading QSO #{qso_id} to HamQTH...')

            state = 'N'

            try:
                logbook.upload_log(self.settings.value(f'callbook/HamQTH_user', ''),
                                   self.settings_form.callbookPassword(CallBookType.HamQTH),
                                   adif_doc)

                self.log.debug(f'Uploaded log to HamQTH')
                state = 'Y'
            except LoginException:
                QtWidgets.QMessageBox.warning(self, self.tr('Upload log error'),
                                              self.tr('Login to HamQTH failed for user') + ': ' + self.settings.value(
                                                  f'callbook/HamQTH_user', ''))
                break
            except QSORejectedException:
                state = 'Y'
                self.log.info('Log rejected, already exists')
            except MissingADIFFieldException as exc:
                self.log.warning(f'A field is missing in QSO #{qso_id} for log upload to HamQTH: "{exc.args[0]}"')
            except CommunicationException as exc:
                QtWidgets.QMessageBox.warning(self, self.tr('Upload log error'),
                                              self.tr('An error occured on uploading to HamQTH') + f':\n"{exc}"')
                break
            finally:
                self.updateQSOField('hamqth', qso_id, state)
        self.refreshTableView(False)

    def selectCallbook(self):
        """Select a different callbook via dialog"""
        # Check if a logbook is available
        if not self.__db_con__.isOpen():
            QtWidgets.QMessageBox.warning(self, self.tr('Select callbook'),
                                          self.tr('There is no database opened for QSO logging.\n'
                                                  'Please open the QSO database first.'))
            self.selectDB()
            if not self.__db_con__.isOpen():
                QtWidgets.QMessageBox.warning(self, self.tr('Select callbook'),
                                              self.tr('No database opened for QSO logging'))
                return

        # File dialog
        res = QtWidgets.QFileDialog.getSaveFileName(
            self,
            self.tr('Select callbook'),
            self.settings.value('lastCallbookPath', os.path.expanduser('~/callbook.db')),
            self.tr('Callbook DB (*.db);;All Files (*.*)'),
            options=QtWidgets.QFileDialog.Option.DontConfirmOverwrite)

        if res[0]:
            try:
                self.__local_cb__ = LocalCallbook(res[0], self.log, self.settings_form.csv_delimiter)
                self.settings.setValue('lastCallbookPath', res[0])

                # Check if DB needs init
                if self.__local_cb__.is_new:
                    if self.__local_cb__.callbook_entries == 0:
                        question = QtWidgets.QMessageBox.question(self, self.tr('Empty callbook'),
                                                                  self.tr(
                                                                      'Should the callbook be initialsed with the existing QSO data?'),
                                                                  QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                                                                  QtWidgets.QMessageBox.StandardButton.Yes)
                        if question == QtWidgets.QMessageBox.StandardButton.Yes:
                            self.initialiseCallbook(True)
                    if self.__local_cb__.history_entries[0] == 0:
                        question = QtWidgets.QMessageBox.question(self, self.tr('Empty call history'),
                                                                  self.tr(
                                                                      'Should the call history be initialsed with the existing contest data?'),
                                                                  QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                                                                  QtWidgets.QMessageBox.StandardButton.Yes)
                        if question == QtWidgets.QMessageBox.StandardButton.Yes:
                            self.initialiseCallHistory(True)
                else:
                    self.updateCallbookStatus()

                self.actionRefresh_callbook.setEnabled(True)
                self.actionExport_callbook.setEnabled(True)
                self.actionImport_callbook.setEnabled(True)
                self.actionRefresh_call_history.setEnabled(True)
                self.actionExport_call_history.setEnabled(True)
                self.actionImport_call_history.setEnabled(True)
            except LocalCallbookDatabaseError:
                QtWidgets.QMessageBox.critical(self,
                                               self.tr('Select callbook'),
                                               self.tr('Selected callbook could not be opend!'))
            except Exception as exc:
                self.log.exception(exc)

    def updateCallbookStatus(self):
        info = (f'{self.__local_cb__.path} '
                f'({self.tr("%n entry", "", self.__local_cb__.callbook_entries)}, '
                f'{self.tr("%n history entry", "", self.__local_cb__.history_entries[0])})')
        self.callbook_status.setText(f'{self.tr("Callbook")}: {"..." + info[-97:] if len(info) > 100 else info}')

    def initialiseCallbook(self, no_warning=False):
        """Initialise an empty or new callbook from existing QSOs"""
        if not no_warning:
            question = QtWidgets.QMessageBox.question(self, self.tr('Refresh callbook'),
                                                      self.tr(
                                                          'Usually the callbook should be initialised and updated automatically with new QSOs.\n'
                                                          'Should the callbook be refreshed anyway?'),
                                                      QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                                                      QtWidgets.QMessageBox.StandardButton.No)
            if question != QtWidgets.QMessageBox.StandardButton.Yes:
                return

        query = self.__db_con__.exec('SELECT call_sign,name,qth,locator,band FROM qsos '
                                     'WHERE band != "11m" '
                                     'ORDER BY date_time')
        if query.lastError().text():
            raise Exception(query.lastError().text())

        while query.next():
            lcd = LocalCallbookData(name=query.value(1), qth=query.value(2), locator=query.value(3))
            self.__local_cb__.update_entry(query.value(0), lcd)

        self.updateCallbookStatus()

    def initialiseCallHistory(self, no_warning=False):
        """Initialise an empty call history from existing contests"""
        if not no_warning:
            question = QtWidgets.QMessageBox.question(self, self.tr('Refresh call history'),
                                                      self.tr(
                                                          'Usually the call history should be initialised and updated automatically with new QSOs.\n'
                                                          'Should the call history be refreshed anyway?'),
                                                      QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                                                      QtWidgets.QMessageBox.StandardButton.No)
            if question != QtWidgets.QMessageBox.StandardButton.Yes:
                return

        query = self.__db_con__.exec('SELECT contest_id, call_sign, crx_data FROM qsos '
                                     'WHERE band != "11m" and contest_id != ""'
                                     'ORDER BY date_time')
        if query.lastError().text():
            raise Exception(query.lastError().text())

        while query.next():
            contest = CONTESTS.get(query.value(0), None)
            if contest:
                excd: ExchangeData = contest.extract_exchange(query.value(2))
                if excd:
                    chd = CallHistoryData(locator=excd.locator, power_class=excd.power, darc_dok=excd.darc_dok,
                                          itu_zone=excd.itu_zone)
                    self.__local_cb__.update_history(query.value(0), query.value(1), chd)

        self.updateCallbookStatus()

    def exportCallbook(self):
        res = QtWidgets.QFileDialog.getSaveFileName(
            self,
            self.tr('Select callbook export file'),
            os.path.join(self.settings.value('lastCallbookExpDir',
                                             self.settings.value('lastExportDir', os.path.abspath(os.curdir))),
                         datetime.date.today().strftime('%Y-%m-%d Callbook.csv')),
            self.tr('CSV-File (*.csv)'))

        if res[0]:
            try:
                self.__local_cb__.export_callbook(res[0])
                self.settings.value('lastCallbookExpDir', os.path.dirname(res[0]))
            except LocalCallbookExportError as exc:
                self.log.error(str(exc))
                QtWidgets.QMessageBox.critical(self,
                                               self.tr('Callbook export'),
                                               self.tr('Error exporting callbook') + f':\n{str(exc)}')

    def importCallbook(self):
        res = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr('Select callbook import file'),
            self.settings.value('lastCallbookImpDir',
                                self.settings.value('lastCallbookExpDir', os.path.abspath(os.curdir))),
            self.tr('CSV-File (*.csv *.txt)'))

        if res[0]:
            try:
                self.__local_cb__.import_callbook(res[0])
                self.settings.setValue('lastCallbookImpDir', os.path.dirname(res[0]))
            except LocalCallbookImportError as exc:
                self.log.error(str(exc))
                QtWidgets.QMessageBox.critical(self,
                                               self.tr('Callbook import'),
                                               self.tr('Error importing callbook') + f':\n{str(exc)}')
            self.updateCallbookStatus()

    def exportCallHistory(self):
        res = QtWidgets.QFileDialog.getSaveFileName(
            self,
            self.tr('Select history export file'),
            os.path.join(self.settings.value('lastHistoryExpDir',
                                             self.settings.value('lastExportDir', os.path.abspath(os.curdir))),
                         datetime.date.today().strftime('%Y-%m-%d Call-History.csv')),
            self.tr('CSV-File (*.csv)'))

        if res[0]:
            try:
                self.__local_cb__.export_history(res[0])
                self.settings.value('lastHistoryExpDir', os.path.dirname(res[0]))
            except LocalCallbookExportError as exc:
                self.log.error(str(exc))
                QtWidgets.QMessageBox.critical(self,
                                               self.tr('Call history export'),
                                               self.tr('Error exporting call history') + f':\n{str(exc)}')

    def importCallHistory(self):
        res = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr('Select history import file'),
            self.settings.value('lastHistoryImpDir',
                                self.settings.value('lastHistoryExpDir', os.path.abspath(os.curdir))),
            self.tr('CSV-File (*.csv *.txt)'))

        if res[0]:
            try:
                self.__local_cb__.import_history(res[0])
                self.settings.setValue('lastHistoryImpDir', os.path.dirname(res[0]))
            except LocalCallbookImportError as exc:
                self.log.error(str(exc))
                QtWidgets.QMessageBox.critical(self,
                                               self.tr('Call history import'),
                                               self.tr('Error importing call history') + f':\n{str(exc)}')
            self.updateCallbookStatus()

    def updateQSOField(self, field: str, qso_id: int, value: str):
        """Update a single QSO field
        :param field: the SQLite column name
        :param qso_id: the QSO Id
        :param value: the new content"""

        if field in self.__sql_cols__:
            self.log.debug(f'Updating {field} for QSO #{qso_id} to {value}...')
            query = QtSql.QSqlQuery(self.__db_con__)
            query.prepare(f'UPDATE qsos SET {field}=? WHERE id = ?')
            query.bindValue(0, value)
            query.bindValue(1, qso_id)
            query.exec()
            if query.lastError().text():
                self.log.error(query.lastError().text())
            self.__db_con__.commit()

    def markQSO(self):
        actions = {
            self.actionMark_QSL_sent: ('qsl_sent', 'Y'),
            self.actionUnmark_QSL_sent: ('qsl_sent', 'N'),
            self.actionMark_QSL_received: ('qsl_rcvd', 'Y'),
            self.actionUnmark_QSL_received: ('qsl_rcvd', 'N'),
            self.actionMark_HamQTH_uploaded: ('hamqth', 'Y'),
        }

        # noinspection PyTypeChecker
        action: QtGui.QAction = self.sender()

        if action not in actions:
            self.log.debug(f'Could not determin action "{action.objectName()}"')
            return

        question = QtWidgets.QMessageBox.question(self, action.text(),
                                                  self.tr('Do you really want to change all selected QSOs?'),
                                                  QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                                                  QtWidgets.QMessageBox.StandardButton.No)

        if question == QtWidgets.QMessageBox.StandardButton.Yes:
            field, state = actions[action]
            for i in self.selectedQSOIds():
                self.updateQSOField(field, i, state)
            self.refreshTableView(sort=False)

    def logImport(self):
        if not self.__db_con__.isOpen():
            self.selectDB()
            if not self.__db_con__.isOpen():
                QtWidgets.QMessageBox.warning(self, self.tr('Import'),
                                              self.tr('No database opened for import'))
                return

        imp_formats = {
            self.tr('ADIF 3 (*.adi *.adif *.zip)'): self.logImportADIF,
            self.tr('ADIF 3 XML (*.adx)'): self.logImportADIF,
            self.tr('CSV-File (*.csv)'): self.logImportCSV,
        }

        if OPTION_OPENPYXL:
            imp_formats[self.tr('Excel-File (*.xlsx)')] = self.logImportExcel

        res = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr('Select import file'),
            self.settings.value('lastImportDir', os.path.abspath(os.curdir)),
            ';;'.join(imp_formats.keys())
        )

        if res[0]:
            # noinspection PyArgumentList
            imp_formats[res[1]](res[0])

            self.settings.setValue('lastImportDir', os.path.dirname(res[0]))

    def logImportExcel(self, file):
        self.log.info('Importing from XLSX...')

        xl_wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        xl_ws = xl_wb.active

        lines = 0

        for ln, row in enumerate(xl_ws.iter_rows(), 1):
            if ln == 1:  # Skip header
                continue

            row = list(row)

            if not len(row) >= len(self.__sql_cols__):
                self.log.warning(f'XLSX import, row {ln} has too few columns.\nSkipped row.')
                QtWidgets.QMessageBox.warning(
                    self,
                    self.tr('Log import XLSX'),
                    f'Row {ln} has too few columns.\nSkipped row.'
                )
                continue

            if not row[1].value or not row[4].value:
                self.log.warning(f'XLSX import, QSO date/time or callsign missing in row {ln}.\nSkipped row.')
                QtWidgets.QMessageBox.warning(
                    self,
                    self.tr('Log import XLSX'),
                    f'QSO date/time or callsign missing in row {ln}.\nSkipped row.'
                )
                continue

            if self.findQSO(row[1].value, row[4].value):
                self.log.info(
                    f'XLSX import, QSO row {ln} already exists for {row[1].value} and {row[4].value}. Skipping row.')
                continue

            if not ln % 100:
                self.log.info(f'XLSX import, processed rows {ln}. Continuing...')

            query = QtSql.QSqlQuery(self.__db_con__)
            query.prepare(self.__db_insert_stmnt__)

            for i, cell in enumerate(row[1:]):
                query.bindValue(i, cell.value)
            query.exec()
            if query.lastError().text():
                QtWidgets.QMessageBox.warning(
                    self,
                    self.tr('Log import XLSX'),
                    f'Row {ln} import error ("{query.lastError().text()}").\nSkipped row.'
                )
            else:
                lines += 1
                self.__db_con__.commit()

        self.refreshTableView()
        self.log.info(f'Imported {lines} QSOs from "{file}"')

    def logImportCSV(self, file):
        self.log.info('Importing from CSV...')

        with open(file, newline='', encoding='utf-8') as cf:
            reader = csv.reader(cf, delimiter=self.settings_form.csv_delimiter)
            lines = 0

            for ln, row in enumerate(reader, 1):
                if ln == 1:  # Skip header
                    continue

                if not len(row) >= len(self.__sql_cols__):
                    self.log.warning(f'CSV import, row {ln} has too few columns.\nSkipped row.')
                    QtWidgets.QMessageBox.warning(
                        self,
                        self.tr('Log import CSV'),
                        f'Row {ln} has too few columns.\nSkipped row.'
                    )

                if not row[1] or not row[4]:
                    self.log.warning(f'CSV import, QSO date/time or callsign missing in row {ln}.\nSkipped row.')
                    QtWidgets.QMessageBox.warning(
                        self,
                        self.tr('Log import CSV'),
                        f'QSO date/time or callsign missing in row {ln}.\nSkipped row.'
                    )
                    continue

                if self.findQSO(row[1], row[4]):
                    self.log.info(
                        f'CSV import, QSO row {ln} already exists for {row[1]} and {row[4]}. Skipping row.')
                    continue

                if not ln % 100:
                    self.log.info(f'CSV import, processed rows {ln}. Continuing...')

                query = QtSql.QSqlQuery(self.__db_con__)
                query.prepare(self.__db_insert_stmnt__)

                for i, val in enumerate(row[1:]):
                    query.bindValue(i, val)
                query.exec()
                if query.lastError().text():
                    QtWidgets.QMessageBox.warning(
                        self,
                        self.tr('Log import CSV'),
                        f'Row {ln} import error ("{query.lastError().text()}").\nSkipped row.'
                    )
                else:
                    lines += 1
                    self.__db_con__.commit()

        self.refreshTableView()
        self.log.info(f'Imported {lines} QSOs from "{file}"')

    def logImportADIF(self, file):
        self.log.info('Importing from ADIF...')

        is_adx: bool = os.path.splitext(file)[-1] == '.adx'
        adi_src_type = ADISourceType.Other
        if is_adx:
            records: list = adx.load(file)['RECORDS']
        else:
            if os.path.splitext(file)[-1] == '.zip':
                with zipfile.ZipFile(file) as z_file:
                    adi_fname = ''
                    for f in z_file.filelist:
                        if os.path.splitext(f.filename)[-1] == '.adi':
                            adi_fname = f.filename
                            break
                    if adi_fname:
                        with z_file.open(adi_fname) as a_file:
                            adi_doc = adi.loads(a_file.read().decode())
                    else:
                        self.log.error(f'ADIF import, no .adi file found in "{file}"')
                        QtWidgets.QMessageBox.critical(
                            self,
                            self.tr('Log import ADIF'),
                            f'ADIF import, no .adi file found in "{file}"'
                        )
                        return
            else:
                adi_doc = adi.load(file)

            if 'HEADER' in adi_doc and 'PROGRAMID' in adi_doc['HEADER']:
                adi_src_type = eval_adi_type(adi_doc['HEADER']['PROGRAMID'].strip())

            records: list = adi_doc['RECORDS']

        imported = 0
        i: int
        r: dict
        for i, r in enumerate(records, 1):
            # Fix ADI data
            if adi_src_type == ADISourceType.eQSLInbox:
                self.log.debug('Fixing eQSL inbox data')
                if 'QSL_SENT' in r:  # Wrong tag, missing tag
                    r['EQSL_QSL_RCVD'] = r.pop('QSL_SENT')
                if 'QSLMSG' in r:  # Not the own message
                    r.pop('QSLMSG')
                if 'RST_SENT' in r:  # Wrong direction
                    r['RST_RCVD'] = r.pop('RST_SENT')
            elif adi_src_type == ADISourceType.eQSLOutbox:
                self.log.debug('Fixing eQSL outbox data')
                if 'QSL_SENT' in r:  # Wrong tag
                    r['EQSL_QSL_SENT'] = r.pop('QSL_SENT')
                if 'QSL_SENT_VIA' in r:  # Wrong usage
                    r.pop('QSL_SENT_VIA')
            elif adi_src_type == ADISourceType.LoTW:
                self.log.debug('Fixing LoTW data')
                if 'QSL_RCVD' in r:  # Wrong tag
                    r['LOTW_QSL_RCVD'] = r.pop('QSL_RCVD')
                r['LOTW_QSL_SENT'] = 'Y'
            elif adi_src_type == ADISourceType.DCL:
                if 'QSL_RCVD' in r:  # Wrong tag
                    r.pop('QSL_RCVD')

            if 'QSO_DATE' not in r or 'TIME_ON' not in r:
                self.log.warning(f'ADIF import, QSO date/time missing in record {i}.\nSkipped record.')
                QtWidgets.QMessageBox.warning(
                    self,
                    self.tr('Log import ADIF'),
                    f'QSO date/time missing in record {i}.\nSkipped record.'
                )
                continue

            if 'CALL' not in r:
                self.log.warning(f'ADIF import, callsign missing in record {i}.\nSkipped record.')
                QtWidgets.QMessageBox.warning(
                    self,
                    self.tr('Log import ADIF'),
                    f'Callsign missing in record {i}.\nSkipped record.'
                )
                continue

            adi_time = r['TIME_ON']
            adi_date = r['QSO_DATE']
            time = f'{adi_time[:2]}:{adi_time[2:4]}' if len(
                adi_time) == 4 else f'{adi_time[:2]}:{adi_time[2:4]}:{adi_time[4:6]}'
            timestamp = f'{adi_date[:4]}-{adi_date[4:6]}-{adi_date[6:8]} {time}'

            if not i % 100:
                self.log.info(f'ADIF import, processed records {i}. Continuing...')

            query = QtSql.QSqlQuery(self.__db_con__)
            qso_values: list = self.findQSO(timestamp, r['CALL'])
            if qso_values:
                self.log.info(f'ADIF import, updating QSO {qso_values[0]} for {timestamp} and {r["CALL"]}...')
                qso_values.append(qso_values.pop(0))
                query.prepare(self.__db_update_stmnt__)
            else:
                query.prepare(self.__db_insert_stmnt__)

            for j, val in enumerate(self._build_adif_import_(r,
                                                             bool(self.settings.value('imp_exp/use_id_adif', 0)),
                                                             bool(self.settings.value('imp_exp/use_station_adif', 0)),
                                                             qso_values)):
                query.bindValue(j, val)
            query.exec()
            if query.lastError().text():
                QtWidgets.QMessageBox.warning(
                    self,
                    self.tr('Log import ADIF'),
                    f'Record {i} import error ("{query.lastError().text()}").\nSkipped record.'
                )
            else:
                imported += 1
                self.__db_con__.commit()

        self.refreshTableView()

        self.log.info(f'Imported {imported} QSOs from "{file}"')

    def _build_adif_import_(self, r, use_cfg_id=False, use_cfg_station=False, update: list = None):
        if update:
            values = update
        else:
            values = [''] * (len(self.__sql_cols__) - 1)
            values[self.__sql_cols__.index('qsl_sent') - 1] = 'N'
            values[self.__sql_cols__.index('qsl_rcvd') - 1] = 'N'
            values[self.__sql_cols__.index('eqsl_sent') - 1] = 'N'
            values[self.__sql_cols__.index('eqsl_rcvd') - 1] = 'N'
            values[self.__sql_cols__.index('lotw_sent') - 1] = 'N'
            values[self.__sql_cols__.index('lotw_rcvd') - 1] = 'N'

        date = r['QSO_DATE']
        timex = r['TIME_ON']
        time = f'{timex[:2]}:{timex[2:4]}' if len(timex) == 4 else f'{timex[:2]}:{timex[2:4]}:{timex[4:6]}'
        values[0] = f'{date[:4]}-{date[4:6]}-{date[6:8]} {time}'
        if 'TIME_OFF' in r:
            date_off = r['QSO_DATE_OFF'] if 'QSO_DATE_OFF' in r else date  # Fallback
            timex_off = r['TIME_OFF']
            time_off = f'{timex_off[:2]}:{timex_off[2:4]}' if len(
                timex_off) == 4 else f'{timex_off[:2]}:{timex_off[2:4]}:{timex_off[4:6]}'
            values[1] = f'{date_off[:4]}-{date_off[4:6]}-{date_off[6:8]} {time_off}'

        if use_cfg_id:
            if not values[self.__sql_cols__.index('own_callsign') - 1]:
                values[self.__sql_cols__.index('own_callsign') - 1] = self.settings.value('station/callSign', '')
            if not values[self.__sql_cols__.index('own_name') - 1]:
                values[self.__sql_cols__.index('own_name') - 1] = self.settings.value('station/name', '')

        if use_cfg_station:
            if not values[self.__sql_cols__.index('radio') - 1]:
                values[self.__sql_cols__.index('radio') - 1] = self.settings.value('station/radio', '')
            if not values[self.__sql_cols__.index('antenna') - 1]:
                values[self.__sql_cols__.index('antenna') - 1] = self.settings.value('station/antenna', '')
            if not values[self.__sql_cols__.index('own_qth') - 1]:
                values[self.__sql_cols__.index('own_qth') - 1] = self.settings.value('station/QTH', '')
            if not values[self.__sql_cols__.index('own_locator') - 1]:
                values[self.__sql_cols__.index('own_locator') - 1] = self.settings.value('station/own_locator', '')

        if 'DISTANCE' not in r and 'GRIDSQUARE' in r and 'MY_GRIDSQUARE' in r:
            # noinspection PyBroadException
            try:
                r['DISTANCE'] = distance(r['GRIDSQUARE'], r['MY_GRIDSQUARE'])
            except Exception:
                pass

        values[self.__sql_cols__.index('channel') - 1] = '-'

        submode_eval = False

        for p in r:
            match p:
                case 'QSO_DATE' | 'TIME_ON' | 'QSO_DATE_OFF' | 'TIME_OFF' | 'APP_DRAGONLOG_CBQSO':
                    continue
                case 'BAND':
                    values[self.__adx_cols__.index(p)] = r[p].lower()
                case 'FREQ':
                    freq = r[p].replace(',', '.')  # Workaround for wrong export from fldigi
                    values[self.__adx_cols__.index(p)] = str(float(freq) * 1000)
                case 'MODE':
                    if r[p] in self.modes['AFU'] or r[p] in self.modes['CB']:  # If it is a main mode
                        values[self.__adx_cols__.index(p)] = r[p]
                    else:  # Workaround for wrong mode export from fldigi
                        for m in self.modes['AFU']:
                            if r[p] in self.modes['AFU'][m]:
                                values[self.__adx_cols__.index(p)] = m
                                values[self.__adx_cols__.index('SUBMODE')] = r[p]
                                submode_eval = True
                                break
                case 'SUBMODE':
                    if not submode_eval:  # Set only if not already evaluated from mode
                        values[self.__adx_cols__.index(p)] = r[p]
                case 'APP':  # Only for ADX as ADI App fields are recognised the standard way
                    for af in r[p]:
                        af_param = f'APP_{af["@PROGRAMID"].upper()}_{af["@FIELDNAME"].upper()}'
                        if af_param in self.__adx_cols__:
                            values[self.__adx_cols__.index(af_param)] = af['$']
                        elif af_param == 'APP_DRAGONLOG_CBCALL':
                            values[self.__adx_cols__.index('CALL')] = af['$']
                        elif af_param == 'APP_DRAGONLOG_CBOWNCALL':
                            values[self.__adx_cols__.index('STATION_CALLSIGN')] = af['$']
                        elif af_param == 'APP_DRAGONLOG_CBQSO' and af['$'] == 'Y':
                            values[self.__adx_cols__.index('BAND')] = '11m'
                case 'OPERATOR':
                    if not 'STATION_CALLSIGN' in r or not r['STATION_CALLSIGN']:
                        values[self.__adx_cols__.index('STATION_CALLSIGN')] = r[p]
                case 'GUEST_OP':
                    if (not 'STATION_CALLSIGN' in r or not r['STATION_CALLSIGN']) and \
                            (not 'OPERATOR' in r or not r['OPERATOR']):
                        values[self.__adx_cols__.index('STATION_CALLSIGN')] = r[p]
                case 'CONTEST_ID':
                    # Try to map ADIF ID to internal
                    values[self.__sql_cols__.index('contest_id') - 1] = CONTEST_ADIF_DL.get(r[p], r[p])
                case p if p in self.__adx_cols__:
                    values[self.__adx_cols__.index(p)] = r[p]
                case p if p + '_INTL' not in r:  # Take non *_INTL only if no suiting *_INTL are in import
                    if p + '_INTL' in self.__adx_cols__:
                        values[self.__adx_cols__.index(p + '_INTL')] = r[p]

        return values

    def addQSOToCallbook(self, qso: dict):
        if not self.__local_cb__:
            return

        call = qso.get('CALL', '')
        if call:
            lcd = LocalCallbookData(name=qso.get('NAME_INTL', qso.get('NAME', '')),
                                    qth=qso.get('QTH_INTL', qso.get('QTH', '')),
                                    locator=qso.get('GRIDSQUARE', ''))
            self.__local_cb__.update_entry(call, lcd)

            contest_id = qso.get('CONTEST_ID', '')
            if contest_id and qso.get('SRX_STRING', ''):
                contest = CONTESTS.get(contest_id, None)
                if contest:
                    excd = contest.extract_exchange(qso.get('SRX_STRING', ''))
                    if excd:
                        chd = CallHistoryData(locator=excd.locator, power_class=excd.power, darc_dok=excd.darc_dok,
                                              itu_zone=excd.itu_zone, rda_number=excd.rda_number)
                        self.__local_cb__.update_history(contest_id, call, chd)

    def fetchCCQSO(self):
        """Fetch QSOs from HamCC"""
        if not self.__db_con__.isOpen():
            self.selectDB()
            if not self.__db_con__.isOpen():
                QtWidgets.QMessageBox.warning(self, self.tr('Saving QSO'),
                                              self.tr('No database opened for saving'))
                return

        added = False
        while self.cc_widget.hasQSOs():
            self.log.info(f'Adding QSO from CassiopeiaConsole to logbook...')

            query = QtSql.QSqlQuery(self.__db_con__)
            query.prepare(self.__db_insert_stmnt__)

            qso = self.cc_widget.retrieveQSO()
            for j, val in enumerate(self._build_adif_import_(qso,
                                                             False,
                                                             bool(self.settings.value('imp_exp/use_station_hamcc', 0)),
                                                             )):
                query.bindValue(j, val)
            query.exec()
            if query.lastError().text():
                self.log.error(
                    f'Import error from watched file ("{query.lastError().text()}").'
                    'Skipped.')

            self.__db_con__.commit()
            added = True
            self.addQSOToCallbook(qso)

        if added:
            self.refreshTableView()
            self.qsoAdded.emit()

    def watchFile(self):
        with open(self.watchFileName, encoding='ASCII') as adi_f:
            adi_str = adi_f.read()

        added = False

        for i, rec in enumerate(adi.loadi(adi_str, self.watchPos)):
            if i == 0:
                continue

            self.watchPos += 1

            if 'QSO_DATE' not in rec or 'TIME_ON' not in rec or 'CALL' not in rec:
                self.log.warning(
                    f'QSO date, time or call missing in record #{self.watchPos + 1} from watched file. Skipped.')
                continue

            adi_time = rec['TIME_ON']
            adi_date = rec['QSO_DATE']
            time = f'{adi_time[:2]}:{adi_time[2:4]}' if len(
                adi_time) == 4 else f'{adi_time[:2]}:{adi_time[2:4]}:{adi_time[4:6]}'
            timestamp = f'{adi_date[:4]}-{adi_date[4:6]}-{adi_date[6:8]} {time}'

            if not self.findQSO(timestamp, rec['CALL']):
                self.log.info(f'Adding QSO #{i} from "{self.watchFileName}" to logbook...')

                query = QtSql.QSqlQuery(self.__db_con__)
                query.prepare(self.__db_insert_stmnt__)

                for j, val in enumerate(self._build_adif_import_(rec,
                                                                 bool(self.settings.value('imp_exp/use_id_watch',
                                                                                          0)),
                                                                 bool(self.settings.value(
                                                                     'imp_exp/use_station_watch', 0)))):
                    query.bindValue(j, val)
                query.exec()
                if query.lastError().text():
                    self.log.error(
                        f'Record #{self.watchPos + 1} import error from watched file ("{query.lastError().text()}").'
                        'Skipped.')

                self.__db_con__.commit()
                added = True

        if added:
            self.refreshTableView()

    def findQSO(self, timestamp, call, minute_range: int = 1) -> list:
        """Find a QSO and return data set as list
        The timestamp is searched for +/- 1 min
        :param timestamp: the timestamp in format "YYYY-MM-DD hh:mm[:ss]"
        :param call: the callsign of the QSO partner
        :param minute_range: the range of minutes +x/-x to search for (minimum 1)
        :return: QSO data as list
        """
        minute_range = minute_range if type(minute_range) is int and minute_range > 1 else 1
        query = self.__db_con__.exec(f'''SELECT * FROM qsos 
            WHERE datetime(date_time) > datetime("{timestamp}", "-{minute_range} minute") 
                AND datetime(date_time) < datetime("{timestamp}", "+{minute_range} minute") 
                AND call_sign = "{call}"''')
        if query.lastError().text():
            raise Exception(query.lastError().text())

        row = []
        if query.next():
            for i in range(len(self.__sql_cols__)):
                row.append(query.value(i))

        return row

    def workedBefore(self, call: str) -> dict[str, dict[str, str]]:
        query = self.__db_con__.exec(f'SELECT call_sign, date_time, contest_id, event from qsos '
                                     f'where call_sign LIKE"%{call}%" ORDER BY date_time DESC')
        if query.lastError().text():
            raise Exception(query.lastError().text())

        worked = {}
        while query.next():
            call = query.value(0)
            if not call in worked:
                worked[call] = {
                    'date_time': query.value(1),
                    'event': query.value(2) if query.value(2) else query.value(3)
                }

        return worked

    def ctrlWatching(self, start):
        if start:
            if not self.__db_con__.isOpen():
                self.selectDB()
                if not self.__db_con__.isOpen():
                    QtWidgets.QMessageBox.warning(self, self.tr('Watching file'),
                                                  self.tr('No database opened for watching files'))
                    return

            app_res = self.watchAppSelect.exec()
            if app_res:
                res = QtWidgets.QFileDialog.getOpenFileName(
                    self,
                    self.tr('Select file to watch'),
                    app_res[1],
                    self.tr('ADIF 3 (*.adi *.adif)')
                )

                if res[0]:
                    if app_res[0] == 'Other':
                        self.settings.setValue('lastWatchFile', res[0])
                    else:
                        self.settings.setValue(f'lastWatchFile{app_res[0]}', res[0])
                    self.watchFileName = res[0]
                    self.watchPos = 0
                    self.watchFile()  # Read file once before looping
                    self.watchTimer.start(2000)
                    self.watch_status.setText(self.tr('Watching file') + f': {os.path.basename(res[0])}')
                    self.actionWatch_file_for_QSOs.setChecked(True)
                    self.actionWatch_file_for_QSOs_TB.setChecked(True)
                    return
        else:
            self.watchTimer.stop()
            self.watch_status.setText(self.tr('Watching file') + ': ' + self.tr('inactiv'))

        self.actionWatch_file_for_QSOs.setChecked(False)
        self.actionWatch_file_for_QSOs_TB.setChecked(False)

    def rigStatusChanged(self, state: bool):
        if state:
            self.hamlib_status.setText(self.tr('Hamlib') + ': ' + self.tr('activ'))
        else:
            self.hamlib_status.setText(self.tr('Hamlib') + ': ' + self.tr('inactiv'))
            self.actionStart_hamlib_TB.setChecked(False)

    def showCC(self):
        if not self.__db_con__.isOpen():
            self.selectDB()
            if not self.__db_con__.isOpen():
                QtWidgets.QMessageBox.warning(self, self.tr('Log QSO'),
                                              self.tr('No database opened for logging'))
                return

        self.ccDockWidget.show()
        self.cc_widget.inputLineEdit.setFocus()

    def exportContestLog(self):
        if not self.__db_con__.isOpen():
            self.selectDB()
            if not self.__db_con__.isOpen():
                QtWidgets.QMessageBox.warning(self, self.tr('Contest Export'),
                                              self.tr('No database opened for contest export'))
                return

        query = self.__db_con__.exec('SELECT DISTINCT contest_id FROM qsos')
        if query.lastError().text():
            raise Exception(query.lastError().text())

        contests = []
        while query.next():
            if query.value(0) in CONTEST_NAMES:
                contests.append(query.value(0))

        if contests:
            contest_dlg = ContestDialog(self, self, self.settings, self.log, contests, self.__cty__)
            contest_dlg.exec()
        else:
            QtWidgets.QMessageBox.warning(self, self.tr('Contest Export'),
                                          self.tr('No contest data available for export'))

    def useFont(self):
        font = self.settings.value('ui/font', 'Inter SlashedZero')
        fontsize = int(self.settings.value('ui/font_size', 9))
        form_fontsize = int(self.settings.value('ui/form_font_size', 10))
        cc_fontsize = int(self.settings.value('ui/cc_font_size', 12))
        css = f'''
            *{{font-family: {font}; font-size: {fontsize}pt}}
            *[cssClass~="QSOFormInput"]{{font-family: {font}; font-size: {form_fontsize}pt}}
            *[cssClass~="CCInput"]{{font-family: {font}; font-size: {cc_fontsize}pt}}
        '''
        self.setStyleSheet(css)
        self.log.info(f'Set font "{font}" to {fontsize} pt')

        if self.qso_form:
            self.qso_form.clear()
        if self.dxspots_widget:
            self.dxspots_widget.tableView.resizeColumnsToContents()
        if self.cstats_widget:
            self.cstats_widget.tableView.resizeColumnsToContents()

        self.QSOTableView.resizeColumnsToContents()

    def cty_load(self, cty_path: str):
        # noinspection PyBroadException
        try:
            self.__cty__ = CountryData(cty_path)
            self.log.debug(f'Using country data from "{cty_path}"')
        except Exception:
            self.__cty__ = CountryData(self.searchFile('data:cty/cty.csv'))
            self.log.debug(f'Using country data default')

    @property
    def cty_version(self) -> str:
        return self.__cty__.version

    @property
    def cty_ver_entity(self) -> str:
        cty_d = self.__cty__.country('VERSION')
        return f'{cty_d.name}, {cty_d.code}'

    def cty_data(self, call: str) -> Country | None:
        try:
            return self.__cty__.country(call)
        except (CountryNotFoundException, CountryCodeNotFoundException):
            pass

    def readQRCode(self):
        if OPTION_QRCODEREADER:
            qr = QSLQRReaderDialog(self, self, self.settings, self.log)
            qr.exec()

    def logImportQRCode(self, record: dict[str, str]):
        """Import single record from QR code"""
        query = QtSql.QSqlQuery(self.__db_con__)
        query.prepare(self.__db_insert_stmnt__)

        for j, val in enumerate(self._build_adif_import_(record,
                                                         bool(self.settings.value('imp_exp/use_id_adif', 0)))):
            query.bindValue(j, val)

        query.exec()
        if query.lastError().text():
            QtWidgets.QMessageBox.warning(
                self,
                self.tr('Log import QSL-QR-Code'),
                f'Record import error ("{query.lastError().text()}").\nSkipped record.'
            )
        else:
            self.__db_con__.commit()
            self.log.info(f'Imported QSO from QSL-QR-Code')
        self.refreshTableView(sort=False)

    def showStatistics(self):
        StatisticsWidget(self, f'{self.programName} - {self.tr("Statistic")}',
                         self.__db_con__, tuple(self.bands.keys()))

    def createHelpDlg(self, title: str, help_text: str):
        help_dialog = QtWidgets.QDialog(self)
        help_dialog.setWindowTitle(f'{self.programName} - {title}')
        help_dialog.setMinimumSize(500, 100)
        help_dialog.setMaximumSize(500, 500)
        help_dialog.setSizePolicy(QtWidgets.QSizePolicy.Policy.Maximum,
                                  QtWidgets.QSizePolicy.Policy.MinimumExpanding)
        verticalLayout = QtWidgets.QVBoxLayout(help_dialog)
        scrollArea = QtWidgets.QScrollArea(help_dialog)
        scrollArea.setWidgetResizable(True)
        scrollAreaWidgetContents = QtWidgets.QWidget()
        verticalLayout_2 = QtWidgets.QVBoxLayout(scrollAreaWidgetContents)
        helpLabel = QtWidgets.QLabel(scrollAreaWidgetContents)
        helpLabel.setTextFormat(QtCore.Qt.TextFormat.MarkdownText)
        helpLabel.setWordWrap(True)
        helpLabel.setOpenExternalLinks(True)
        verticalLayout_2.addWidget(helpLabel)
        scrollArea.setWidget(scrollAreaWidgetContents)
        verticalLayout.addWidget(scrollArea)
        horizontalLayout = QtWidgets.QHBoxLayout()
        horizontalSpacer = QtWidgets.QSpacerItem(40, 20,
                                                 QtWidgets.QSizePolicy.Policy.Expanding,
                                                 QtWidgets.QSizePolicy.Policy.Minimum)
        horizontalLayout.addItem(horizontalSpacer)
        pushButton = QtWidgets.QPushButton(help_dialog)
        pushButton.setText(self.tr('Ok'))
        horizontalLayout.addWidget(pushButton)
        verticalLayout.addLayout(horizontalLayout)
        # noinspection PyUnresolvedReferences
        pushButton.clicked.connect(help_dialog.accept)
        helpLabel.setText(help_text)
        return help_dialog

    # noinspection PyPep8Naming
    def showHelp(self):
        if not self.help_dialog:
            with open(self.searchFile('help:README.md')) as hf:
                help_text = hf.read()
            self.help_dialog = self.createHelpDlg(self.tr("Help"), help_text)
        self.help_dialog.show()

    def showShortcuts(self):
        if not self.help_sc_dialog:
            with open(self.searchFile('help:SHORTCUTS.md')) as hf:
                help_text = hf.read()
            self.help_sc_dialog = self.createHelpDlg(self.tr("Shortcuts"), help_text)
        self.help_sc_dialog.show()

    def showCCHelp(self):
        if not self.help_cc_dialog:
            with open(self.searchFile('help:README_HAMCC.md')) as hf:
                help_text = hf.read()
            self.help_cc_dialog = self.createHelpDlg(self.tr("CassipeiaConsole"), help_text)
        self.help_cc_dialog.show()

    def showContestHelp(self):
        if not self.help_contest_dialog:
            self.help_contest_dialog = self.createHelpDlg(self.tr("Available Contests"), build_contest_list())
        self.help_contest_dialog.show()

    @property
    def programName(self):
        return __prog_name__

    @property
    def programVersion(self):
        return __version_str__

    def showAbout(self):
        cr = sys.copyright.replace('\n\n', '\n')

        opt_xl = f'\n\nOpenPyXL {openpyxl.__version__}: Copyright (c) 2010 openpyxl' if OPTION_OPENPYXL else ''
        opt_qrcode = (f'\nopencv-python {cv2_version}: Copyright (c) Olli-Pekka Heinisuo'
                      f'\npyzbar {pyzbar_version}: Copyright (c) 2016 The Trustees of the Natural History Museum, London') if OPTION_QRCODEREADER else ''

        cty_ver = self.cty_version
        cty_ent = self.cty_ver_entity

        QtWidgets.QMessageBox.about(
            self,
            f'{self.programName} - {self.tr("About")}',
            f'{self.tr("Version")}: {self.programVersion}\n'
            f'{self.tr("Author")}: {__author_name__} <{__author_email__}>\n{__copyright__}'
            f'\n\nPython {sys.version.split()[0]}: {cr}' +
            opt_xl +
            '\nmaidenhead: Copyright (c) 2018 Michael Hirsch, Ph.D.' +
            f'\nPyADIF-File {adif_file.__version_str__}: {adif_file.__copyright__}' +
            f'\nHamCC {hamcc.__version_str__}: {hamcc.__copyright__}' +
            opt_qrcode +
            '\n\nIcons: Crystal Project, Copyright (c) 2006-2007 Everaldo Coelho'
            '\nFlags: Flagpedia.net, https://flagpedia.net'
            '\nFont: Inter, Copyright (c) 2016 The Inter Project Authors (https://github.com/rsms/inter)'
            '\nDragon and QR-Code icon by Icons8 https://icons8.com'
            f'\n\nCountry Data: by AD1C, Copyright (c) since 1994\nVersion: {cty_ver}, Entity: {cty_ent}'
        )

    def showAboutQt(self):
        QtWidgets.QMessageBox.aboutQt(self, self.programName + ' - ' + self.tr('About Qt'))

    def closeEvent(self, e):
        self.log.info(f'Quiting {self.programName}...')
        if self.__db_con__:
            self.log.debug('Reducing and optimising database...')
            self.__db_con__.exec('VACUUM;')
            self.__db_con__.exec('PRAGMA optimize;')
            self.log.debug('Closing database...')
            self.__db_con__.close()
        if self.__local_cb__:
            self.__local_cb__.close()

        self.settings.setValue('ui/show_log', int(self.logDockWidget.isVisible()))
        self.settings.setValue('ui/log_dock_area', self.dockWidgetArea(self.logDockWidget).value)
        self.settings.setValue('ui/log_dock_float', int(self.logDockWidget.isFloating()))

        self.settings.setValue('ui/show_filter', int(self.filterDockWidget.isVisible()))
        self.settings.setValue('ui/filter_dock_area', self.dockWidgetArea(self.filterDockWidget).value)
        self.settings.setValue('ui/filter_dock_float', int(self.filterDockWidget.isFloating()))

        self.settings.setValue('ui/show_qso', int(self.qsoDockWidget.isVisible()))
        self.settings.setValue('ui/qso_dock_area', self.dockWidgetArea(self.qsoDockWidget).value)
        self.settings.setValue('ui/qso_dock_float', int(self.qsoDockWidget.isFloating()))

        self.settings.setValue('ui/show_cc', int(self.ccDockWidget.isVisible()))
        self.settings.setValue('ui/cc_dock_area', self.dockWidgetArea(self.ccDockWidget).value)
        self.settings.setValue('ui/cc_dock_float', int(self.ccDockWidget.isFloating()))

        self.settings.setValue('ui/show_dxspots', int(self.dxSpotsDockWidget.isVisible()))
        self.settings.setValue('ui/dxspots_dock_area', self.dockWidgetArea(self.dxSpotsDockWidget).value)
        self.settings.setValue('ui/dxspots_dock_float', int(self.dxSpotsDockWidget.isFloating()))

        self.settings.setValue('ui/show_contest_stats', int(self.contestStatDockWidget.isVisible()))
        self.settings.setValue('ui/contest_stats_dock_area', self.dockWidgetArea(self.contestStatDockWidget).value)
        self.settings.setValue('ui/contest_stats_dock_float', int(self.contestStatDockWidget.isFloating()))

        self.__rigctl__.ctrlRigctld(False)
        e.accept()


def main():
    app = QtWidgets.QApplication(sys.argv)
    app_path = os.path.dirname(__file__)

    css = '''QToolBox::tab {
        background: #b0c4de;
        border-radius: 2px;
        color: black;
    }'''
    app.setStyleSheet(css)

    translator = QtCore.QTranslator(app)
    translator.load(os.path.abspath(app_path + '/data/i18n/DragonLog_' + QtCore.QLocale.system().name()))
    app.installTranslator(translator)

    file = None
    ini = None
    args = app.arguments()
    while len(args) > 1:
        arg = args.pop(1)
        if arg == '-ini':
            ini = args.pop(1) if len(args) > 1 else None
        else:
            file = arg
            if not (os.path.isfile(file) and os.path.splitext(file)[-1] in ('.qlog', '.sqlite')):
                file = None

    QtCore.QDir.addSearchPath('icons', app_path + '/icons')
    QtCore.QDir.addSearchPath('data', app_path + '/data')
    QtCore.QDir.addSearchPath('help', app_path + '/data')

    dl = DragonLog(file, app_path, ini)
    dl.show()

    sys.exit(app.exec())


# Get old behaviour on printing a traceback on exceptions
sys._excepthook = sys.excepthook


# noinspection PyProtectedMember
def except_hook(cls, exception, traceback):
    # noinspection PyUnresolvedReferences
    sys._excepthook(cls, exception, traceback)
    sys.exit(1)


sys.excepthook = except_hook

if __name__ == '__main__':
    main()
