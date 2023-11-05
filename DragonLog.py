import os
import csv
import sys
import json
import math
import string
import datetime
from xml.etree.ElementTree import ElementTree

from PyQt6 import QtCore, QtWidgets, QtSql, QtGui
import openpyxl
from openpyxl.styles import Font
import maidenhead
import xmlschema
import adif_file

import DragonLog_MainWindow_ui

__prog_name__ = 'DragonLog'
__prog_desc__ = 'Log QSO for Ham radio'
__author_name__ = 'Andreas Schawo'
__author_email__ = 'andreas@schawo.de'
__copyright__ = 'Copyright 2023 by Andreas Schawo,licensed under CC BY-SA 4.0'

import __version__ as version

__version__ = version.__version__

from DragonLog_QSOForm import QSOForm
from DragonLog_Settings import Settings

if version.__branch__:
    __version__ += '-' + version.__branch__
if version.__unclean__:
    __version__ += '-unclean'


class DatabaseOpenException(Exception):
    pass


class DatabaseWriteException(Exception):
    pass


class BackgroundBrushDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, color_map, column):
        super(BackgroundBrushDelegate, self).__init__()

        self.color_map = color_map
        self.column = column

    def getColor(self, value):
        color = [255, 255, 255, 0]  # white as fallback

        if value in self.color_map:
            color = self.color_map[value]
        elif 'default' in self.color_map:
            color = self.color_map['default']

        return color

    def initStyleOption(self, option: QtWidgets.QStyleOptionViewItem, index: QtCore.QModelIndex):
        super().initStyleOption(option, index)

        option.backgroundBrush = QtGui.QBrush(
            QtGui.QColor(*self.getColor(index.model().data(index.siblingAtColumn(self.column)))))


class DragonLog(QtWidgets.QMainWindow, DragonLog_MainWindow_ui.Ui_MainWindow):
    __sql_cols__ = ('id', 'date_time', 'own_callsign', 'call_sign', 'name', 'qth', 'locator',
                    'rst_sent', 'rst_rcvd', 'band', 'mode', 'freq', 'channel', 'power',
                    'own_name', 'own_qth', 'own_locator', 'radio', 'antenna', 'remarks', 'dist')

    __adx_cols__ = ('QSO_DATE/TIME_ON', 'STATION_CALLSIGN', 'CALL', 'NAME_INTL', 'QTH_INTL', 'GRIDSQUARE',
                    'RST_SENT', 'RST_RCVD', 'BAND', 'MODE', 'FREQ', 'APP_DRAGONLOG_CHANNEL', 'TX_PWR',
                    'MY_NAME_INTL', 'MY_CITY_INTL', 'MY_GRIDSQUARE', 'MY_RIG_INTL', 'MY_ANTENNA_INTL', 'NOTES_INTL',
                    'DISTANCE')

    __db_create_stmnt__ = '''CREATE TABLE IF NOT EXISTS "qsos" (
                            "id"    INTEGER PRIMARY KEY NOT NULL,
                            "date_time"   NUMERIC,
                            "own_callsign" TEXT,
                            "call_sign"  TEXT,
                            "name"  TEXT,
                            "qth"    TEXT,
                            "locator" TEXT,
                            "rst_sent" TEXT,
                            "rst_rcvd" TEXT,
                            "band"    TEXT,
                            "mode"   TEXT,
                            "freq"  REAL,
                            "channel"  INTEGER,
                            "power"  REAL,
                            "own_name"  TEXT,
                            "own_qth"   TEXT,
                            "own_locator" TEXT,
                            "radio"   TEXT,
                            "antenna"   TEXT,
                            "remarks"   TEXT,
                            "dist" INTEGER
                        );'''

    __db_insert_stmnt__ = 'INSERT INTO qsos (' + ",".join(__sql_cols__[1:]) + ') ' \
                                                                              'VALUES (' + ",".join(
        ["?"] * (len(__sql_cols__) - 1)) + ')'
    __db_update_stmnt__ = 'UPDATE qsos SET ' + "=?,".join(__sql_cols__[1:]) + '=? ' \
                                                                              'WHERE id = ?'
    __db_select_stmnt__ = f'SELECT * FROM qsos'

    @staticmethod
    def calc_distance(mh_pos1: str, mh_pos2: str):
        # noinspection PyBroadException
        try:
            pos1 = maidenhead.to_location(mh_pos1, True)
            pos2 = maidenhead.to_location(mh_pos2, True)

            mlat = math.radians(pos1[0])
            mlon = math.radians(pos1[1])
            plat = math.radians(pos2[0])
            plon = math.radians(pos2[1])

            return int(6371.01 * math.acos(
                math.sin(mlat) * math.sin(plat) + math.cos(mlat) * math.cos(plat) * math.cos(mlon - plon)))
        except Exception:
            print(f'Exception calcing distance between "{mh_pos1}" amd "{mh_pos2}"')
            return 0

    @staticmethod
    def searchFile(name):
        file = QtCore.QFile(name)
        if file.exists():
            return file.fileName()

    def __init__(self, file=None, app_path='.'):
        super().__init__()

        print(f'Starting {__prog_name__}...')

        self.setupUi(self)

        self.app_path = app_path
        self.help_dialog = None

        self.settings = QtCore.QSettings(self.tr(__prog_name__))

        self.dummy_status = QtWidgets.QLabel()
        self.statusBar().addPermanentWidget(self.dummy_status)
        self.hamlib_status = QtWidgets.QLabel(self.tr('Hamlib') + ': ' + self.tr('inactiv'))
        self.statusBar().addPermanentWidget(self.hamlib_status)
        self.hamlib_error = QtWidgets.QLabel('')
        self.statusBar().addPermanentWidget(self.hamlib_error)

        with open(self.searchFile('data:bands.json')) as bj:
            self.bands: dict = json.load(bj)
        with open(self.searchFile('data:modes.json')) as mj:
            self.modes: dict = json.load(mj)
        with open(self.searchFile('data:cb_channels.json')) as cj:
            self.cb_channels: dict = json.load(cj)

        with open(self.searchFile('data:color_map.json')) as cmj:
            color_map: dict = json.load(cmj)
        self.QSOTableView.setItemDelegate(BackgroundBrushDelegate(color_map, self.__sql_cols__.index('band')))

        self.settings_form = Settings(self, self.settings, self.hamlib_status)

        self.qso_form = QSOForm(self, self.bands, self.modes, self.settings, self.settings_form,
                                self.cb_channels, self.hamlib_error)
        self.keep_logging = False

        self.__headers__ = (
            self.tr('QSO'),
            self.tr('Date/Time'),
            self.tr('Own call sign'),
            self.tr('Call sign'),
            self.tr('Name'),
            self.tr('QTH'),
            self.tr('Locator'),
            self.tr('RST sent'),
            self.tr('RST rcvd'),
            self.tr('Band'),
            self.tr('Mode'),
            self.tr('Frequency'),
            self.tr('Channel'),
            self.tr('Power'),
            self.tr('Own Name'),
            self.tr('Own QTH'),
            self.tr('Own Locator'),
            self.tr('Radio'),
            self.tr('Antenna'),
            self.tr('Remarks'),
            self.tr('Distance'),
        )

        self.__header_map__ = dict(zip(self.__sql_cols__, self.__headers__))

        self.adx_export_schema = xmlschema.XMLSchema(self.searchFile('data:adif/adx314.xsd'))
        self.adx_import_schema = xmlschema.XMLSchema(self.searchFile('data:adif/adx314generic.xsd'))

        self.__db_con__ = QtSql.QSqlDatabase.addDatabase('QSQLITE', 'main')

        if file:
            print(f'Opening database from commandline {file}...')
            self.connectDB(file)
        elif self.settings.value('lastDatabase', None):
            if os.path.isfile(self.settings.value('lastDatabase', None)):
                print(f'Opening last database {self.settings.value("lastDatabase", None)}...')
                self.connectDB(self.settings.value('lastDatabase', None))
            else:
                print(f'Opening last database {self.settings.value("lastDatabase", None)} failed!')

    def showSettings(self):
        self.settings_form.exec()

    def selectDB(self):
        res = QtWidgets.QFileDialog.getSaveFileName(
            self,
            self.tr('Select file'),
            self.settings.value('lastDatabase', None),
            self.tr('QSO-Log (*.qlog);;All Files (*.*)'),
            options=QtWidgets.QFileDialog.Option.DontConfirmOverwrite)

        if res[0]:
            print(f'Selected database {res[0]}')
            self.connectDB(res[0])

    def checkDB(self, db_file):
        # Check database for missing cols
        res = self.__db_con__.exec('SELECT GROUP_CONCAT(NAME,",") as columns FROM PRAGMA_TABLE_INFO("qsos")')
        res.next()
        db_cols = res.value('columns')
        db_cols_l = db_cols.split(',')
        missing_cols = False
        for col in self.__sql_cols__:
            if col not in db_cols_l:
                missing_cols = True
                break

        if missing_cols:
            path, db_name = os.path.split(db_file)
            bck_name = os.path.join(path, datetime.date.today().strftime('%Y-%m-%d') + ' ' + db_name)

            QtWidgets.QMessageBox.warning(self, self.tr('Database structure out-dated'),
                                          self.tr('The database structure is out-dated and needs a conversion\n\n'
                                                  'A backup will be generated:') + f'\n"{bck_name}"',
                                          )

            # Create backup
            self.__db_con__.close()
            os.rename(db_file, bck_name)

            # Open new DB
            self.__db_con__.setDatabaseName(db_file)
            if self.__db_con__.lastError().text():
                raise DatabaseOpenException(self.__db_con__.lastError().text())
            self.__db_con__.open()
            self.__db_con__.exec(self.__db_create_stmnt__)
            if self.__db_con__.lastError().text():
                raise DatabaseOpenException(self.__db_con__.lastError().text())

            # Open backup
            db_con_bck = QtSql.QSqlDatabase.addDatabase('QSQLITE', 'backup')
            db_con_bck.setDatabaseName(bck_name)
            if db_con_bck.lastError().text():
                raise DatabaseOpenException(db_con_bck.lastError().text())
            db_con_bck.open()
            self.__db_con__.exec(self.__db_create_stmnt__)
            if self.__db_con__.lastError().text():
                raise DatabaseOpenException(self.__db_con__.lastError().text())

            q_read = db_con_bck.exec(f'SELECT {db_cols} FROM qsos')
            if q_read.lastError().text():
                raise Exception(q_read.lastError().text())

            insert_stmnt = f'INSERT INTO qsos ({db_cols}) ' \
                           'VALUES (' + ",".join(["?"] * len(db_cols_l)) + ')'

            while q_read.next():
                row = []
                for i in range(len(db_cols_l)):
                    row.append(q_read.value(i))

                q_write = QtSql.QSqlQuery(self.__db_con__)
                q_write.prepare(insert_stmnt)
                for i, val in enumerate(row):
                    q_write.bindValue(i, val)
                q_write.exec()
                if q_write.lastError().text():
                    raise Exception(q_write.lastError().text())

            self.__db_con__.commit()
            db_con_bck.close()

            QtWidgets.QMessageBox.information(self, self.tr('Database conversion'),
                                              self.tr('Database conversion finished')
                                              )

    def connectDB(self, db_file):
        db_file = os.path.abspath(db_file)
        try:
            if self.__db_con__.isOpen():
                self.__db_con__.close()

            self.__db_con__.setDatabaseName(db_file)
            if self.__db_con__.lastError().text():
                raise DatabaseOpenException(self.__db_con__.lastError().text())
            self.__db_con__.open()
            self.__db_con__.exec(self.__db_create_stmnt__)
            if self.__db_con__.lastError().text():
                raise DatabaseOpenException(self.__db_con__.lastError().text())

            self.checkDB(db_file)

            model = QtSql.QSqlTableModel(self, self.__db_con__)
            model.setTable('qsos')

            for c in self.__sql_cols__:
                model.setHeaderData(self.__sql_cols__.index(c),
                                    QtCore.Qt.Orientation.Horizontal,
                                    self.__header_map__[c])

            self.QSOTableView.setModel(model)
            self.QSOTableView.hideColumn(self.__sql_cols__.index('own_name'))
            self.QSOTableView.hideColumn(self.__sql_cols__.index('own_qth'))
            self.QSOTableView.hideColumn(self.__sql_cols__.index('own_locator'))
            self.QSOTableView.hideColumn(self.__sql_cols__.index('radio'))
            self.QSOTableView.hideColumn(self.__sql_cols__.index('antenna'))
            self.QSOTableView.hideColumn(self.__sql_cols__.index('remarks'))
            self.QSOTableView.sortByColumn(1, QtCore.Qt.SortOrder.AscendingOrder)
            self.QSOTableView.resizeColumnsToContents()

            print(f'Opened database {db_file}')
            self.settings.setValue('lastDatabase', db_file)
            self.setWindowTitle(__prog_name__ + ' - ' + db_file)
        except DatabaseOpenException as exc:
            if db_file == self.settings.value('lastDatabase', None):
                self.settings.setValue('lastDatabase', None)

            QtWidgets.QMessageBox.critical(
                self,
                f'{__prog_name__} - {self.tr("Error")}',
                str(exc))

    def logQSO(self):
        self.qso_form.clear()
        if not self.keep_logging:
            self.qso_form.reset()

        dt = QtCore.QDateTime.currentDateTimeUtc()
        self.qso_form.dateEdit.setDate(dt.date())
        self.qso_form.timeEdit.setTime(dt.time())

        if self.qso_form.exec():
            print('Logging QSO...')

            if self.qso_form.autoDateCheckBox.isChecked():
                date_time = QtCore.QDateTime.currentDateTimeUtc().toString('yyyy-MM-dd HH:mm:ss')
            else:
                date_time = self.qso_form.dateEdit.text() + ' ' + self.qso_form.timeEdit.text()

            band = self.qso_form.bandComboBox.currentText()

            values = (
                date_time,
                self.qso_form.ownCallSignLineEdit.text(),
                self.qso_form.callSignLineEdit.text(),
                self.qso_form.nameLineEdit.text(),
                self.qso_form.QTHLineEdit.text(),
                self.qso_form.locatorLineEdit.text(),
                self.qso_form.RSTSentLineEdit.text(),
                self.qso_form.RSTRcvdLineEdit.text(),
                band,
                self.qso_form.modeComboBox.currentText(),
                self.qso_form.freqDoubleSpinBox.value() if self.qso_form.freqDoubleSpinBox.value() >= self.bands[band][
                    0] else '',
                self.qso_form.channelComboBox.currentText() if band == '11m' else '-',
                self.qso_form.powerSpinBox.value() if self.qso_form.powerSpinBox.value() > 0 else '',
                self.qso_form.ownNameLineEdit.text(),
                self.qso_form.ownQTHLineEdit.text(),
                self.qso_form.ownLocatorLineEdit.text(),
                self.qso_form.radioLineEdit.text(),
                self.qso_form.antennaLineEdit.text(),
                self.qso_form.remarksTextEdit.toPlainText().strip(),
                self.calc_distance(self.qso_form.locatorLineEdit.text(), self.qso_form.ownLocatorLineEdit.text())
            )

            query = QtSql.QSqlQuery(self.__db_con__)
            query.prepare(self.__db_insert_stmnt__)
            for i, val in enumerate(values):
                query.bindValue(i, val)
            query.exec()
            if query.lastError().text():
                raise Exception(query.lastError().text())

            self.__db_con__.commit()
            self.QSOTableView.model().select()
            self.QSOTableView.resizeColumnsToContents()
        else:
            print('Logging aborted')
            self.keep_logging = False

        self.hamlib_error.setText('')

    def logMultiQSOs(self):
        self.keep_logging = True

        self.qso_form.setWindowTitle(self.tr('Log multi QSOs'))

        while self.keep_logging:
            self.logQSO()

        self.qso_form.reset()  # To reset window title

    def deleteQSO(self):
        res = QtWidgets.QMessageBox.question(self, self.tr('Delete QSO'),
                                             self.tr('Do you really want to delete the selected QSO(s)?'),
                                             defaultButton=QtWidgets.QMessageBox.StandardButton.No)

        if res == QtWidgets.QMessageBox.StandardButton.Yes:
            done_ids = []

            for i in self.QSOTableView.selectedIndexes():
                qso_id = self.QSOTableView.model().data(i.siblingAtColumn(0))

                if qso_id in done_ids or qso_id is None:
                    continue
                done_ids.append(qso_id)

                print(f'Deleting QSO "{qso_id}"...')
                query = QtSql.QSqlQuery(self.__db_con__)
                query.prepare('DELETE FROM qsos where id == ?')
                query.bindValue(0, qso_id)
                query.exec()

                if query.lastError().text():
                    raise Exception(query.lastError().text())

            self.__db_con__.commit()
            self.QSOTableView.model().select()
            self.QSOTableView.resizeColumnsToContents()

    def changeQSO(self):
        done_ids = []

        for i in self.QSOTableView.selectedIndexes():
            qso_id = self.QSOTableView.model().data(i.siblingAtColumn(0))

            if qso_id in done_ids or qso_id is None:
                continue
            done_ids.append(qso_id)

            self.qso_form.autoDateCheckBox.setChecked(False)
            self.qso_form.stationGroupBox.setChecked(False)
            self.qso_form.identityGroupBox.setChecked(False)
            self.qso_form.autoDateCheckBox.setEnabled(False)
            self.qso_form.stationGroupBox.setCheckable(False)
            self.qso_form.identityGroupBox.setCheckable(False)

            date, time = self.QSOTableView.model().data(i.siblingAtColumn(self.__sql_cols__.index('date_time'))).split()
            self.qso_form.dateEdit.setDate(QtCore.QDate.fromString(date, 'yyyy-MM-dd'))
            self.qso_form.timeEdit.setTime(QtCore.QTime.fromString(time))
            self.qso_form.ownCallSignLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('own_callsign'))))
            self.qso_form.callSignLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('call_sign'))))
            self.qso_form.nameLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('name'))))
            self.qso_form.QTHLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('qth'))))
            self.qso_form.locatorLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('locator'))))
            self.qso_form.RSTSentLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('rst_sent'))))
            self.qso_form.RSTRcvdLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('rst_rcvd'))))

            band = self.QSOTableView.model().data(i.siblingAtColumn(self.__sql_cols__.index('band')))
            self.qso_form.bandComboBox.setCurrentText(band)

            self.qso_form.modeComboBox.setCurrentText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('mode'))))

            try:
                freq = float(self.QSOTableView.model().data(i.siblingAtColumn(self.__sql_cols__.index('freq'))))
            except ValueError:
                freq = self.bands[band][0] - self.bands[band][2]
            self.qso_form.freqDoubleSpinBox.setValue(freq)

            if band == '11m':
                self.qso_form.channelComboBox.setCurrentIndex(-1)
                channel = self.QSOTableView.model().data(i.siblingAtColumn(self.__sql_cols__.index('channel')))
                self.qso_form.channelComboBox.setCurrentText(str(channel) if channel else '-')
            else:
                self.qso_form.channelComboBox.setCurrentIndex(-1)

            try:
                power = int(self.QSOTableView.model().data(i.siblingAtColumn(
                    self.__sql_cols__.index('power'))))
            except ValueError:
                power = 0
            self.qso_form.powerSpinBox.setValue(power)
            self.qso_form.ownNameLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('own_name'))))
            self.qso_form.ownQTHLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('own_qth'))))
            self.qso_form.ownLocatorLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('own_locator'))))
            self.qso_form.radioLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('radio'))))
            self.qso_form.antennaLineEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('antenna'))))
            self.qso_form.remarksTextEdit.setText(self.QSOTableView.model().data(i.siblingAtColumn(
                self.__sql_cols__.index('remarks'))))

            self.qso_form.setWindowTitle(self.tr('Change QSO') + f' #{qso_id}')

            if self.qso_form.exec():
                print(f'Changing QSO {qso_id}...')

                band = self.qso_form.bandComboBox.currentText()

                values = (
                    self.qso_form.dateEdit.text() + ' ' + self.qso_form.timeEdit.text(),
                    self.qso_form.ownCallSignLineEdit.text(),
                    self.qso_form.callSignLineEdit.text(),
                    self.qso_form.nameLineEdit.text(),
                    self.qso_form.QTHLineEdit.text(),
                    self.qso_form.locatorLineEdit.text(),
                    self.qso_form.RSTSentLineEdit.text(),
                    self.qso_form.RSTRcvdLineEdit.text(),
                    band,
                    self.qso_form.modeComboBox.currentText(),
                    self.qso_form.freqDoubleSpinBox.value() if self.qso_form.freqDoubleSpinBox.value() >=
                                                               self.bands[band][0] else '',
                    self.qso_form.channelComboBox.currentText() if band == '11m' else '-',
                    self.qso_form.powerSpinBox.value() if self.qso_form.powerSpinBox.value() > 0 else '',
                    self.qso_form.ownNameLineEdit.text(),
                    self.qso_form.ownQTHLineEdit.text(),
                    self.qso_form.ownLocatorLineEdit.text(),
                    self.qso_form.radioLineEdit.text(),
                    self.qso_form.antennaLineEdit.text(),
                    self.qso_form.remarksTextEdit.toPlainText().strip(),
                    self.calc_distance(self.qso_form.locatorLineEdit.text(), self.qso_form.ownLocatorLineEdit.text()),
                    qso_id,
                )

                query = QtSql.QSqlQuery(self.__db_con__)
                query.prepare(self.__db_update_stmnt__)

                for col, val in enumerate(values):
                    query.bindValue(col, val)
                query.exec()
                if query.lastError().text():
                    raise Exception(query.lastError().text())
            else:
                print(f'Changing QSO(s) aborted')
                break

        self.__db_con__.commit()
        self.QSOTableView.model().select()
        self.QSOTableView.resizeColumnsToContents()

    def export(self):
        exp_formats = {
            self.tr('Excel-File (*.xlsx)'): self.exportExcel,
            self.tr('CSV-File (*.csv)'): self.exportCSV,
            self.tr('ADIF 3 (*.adx *.adi *.adif)'): self.exportADIF,
        }

        res = QtWidgets.QFileDialog.getSaveFileName(
            self,
            self.tr('Select export file'),
            self.settings.value('lastExportDir', os.path.abspath(os.curdir)),
            ';;'.join(exp_formats.keys()))

        if res[0]:
            exp_formats[res[1]](res[0])

            self.settings.setValue('lastExportDir', os.path.abspath(os.path.dirname(res[0])))

    def exportCSV(self, file):
        print('Exporting to CSV...')

        with open(file, 'w', newline='') as cf:
            writer = csv.writer(cf, delimiter=';', dialect=csv.excel)

            # Write header
            writer.writerow(self.__headers__)

            # Write content
            query = self.__db_con__.exec(self.__db_select_stmnt__)
            if query.lastError().text():
                raise Exception(query.lastError().text())

            while query.next():
                row = []
                for i in range(len(self.__sql_cols__)):
                    row.append(query.value(i))
                writer.writerow(row)

    def exportExcel(self, file):
        print('Exporting to Excel...')
        xl_wb = openpyxl.Workbook()
        xl_wb.properties.title = self.tr('Exported QSO log')
        xl_wb.properties.description = f'{__prog_name__} {__version__}'
        xl_wb.properties.creator = os.getlogin()

        xl_ws = xl_wb.active
        xl_ws.title = 'QSOs'
        xl_ws.freeze_panes = 'A2'

        # Write header
        column = 1
        for h in self.__headers__:
            cell = xl_ws.cell(column=column, row=1, value=h)
            cell.font = Font(bold=True)
            column += 1

        # Write content
        query = self.__db_con__.exec(self.__db_select_stmnt__)
        row = 2
        while query.next():
            for col in range(len(self.__headers__)):
                xl_ws.cell(column=col + 1, row=row, value=query.value(col))
            row += 1

        # Set auto filter
        xl_ws.auto_filter.ref = f'A1:{string.ascii_uppercase[len(self.__headers__) - 1]}1'

        # Fit size to content is not available so set fixed approximations
        # Fixme: fix to current column layout
        col_widths = (10, 20, 15, 15, 25, 25, 15, 10, 10, 10, 10, 15, 10, 25, 15, 25, 25, 40, 10)
        for c, w in zip(string.ascii_uppercase[:len(col_widths)], col_widths):
            xl_ws.column_dimensions[c].width = w

        # Finally save
        try:
            xl_wb.save(file)
            print(f'Saved "{file}"')
        except OSError as e:
            QtWidgets.QMessageBox.critical(
                self,
                f'{__prog_name__} - {self.tr("Error")}',
                str(e))

    @staticmethod
    def replaceUmlautsLigatures(text: str):
        text = text.replace('Ä', 'Ae')
        text = text.replace('Ö', 'Oe')
        text = text.replace('Ü', 'Ue')
        text = text.replace('ä', 'ae')
        text = text.replace('ö', 'oe')
        text = text.replace('ü', 'ue')
        text = text.replace('ß', 'ss')
        return text

    def exportADIF(self, file):
        print('Exporting to ADIF...')

        is_adx: bool = os.path.splitext(file)[-1] == '.adx'

        doc = {
            'HEADER':
                {
                    'ADIF_VER': '3.1.4',
                    'PROGRAMID': __prog_name__,
                    'PROGRAMVERSION': __version__,
                    'CREATED_TIMESTAMP': QtCore.QDateTime.currentDateTimeUtc().toString('yyyyMMdd HHmmss')
                },
            'RECORDS': None,
        }

        records = []

        query = self.__db_con__.exec(self.__db_select_stmnt__)
        if query.lastError().text():
            raise Exception(query.lastError().text())

        while query.next():
            band = query.value(self.__sql_cols__.index('band'))

            if band == '11m' and not self.settings.value('station_cb/cb_exp_adif', 0):
                continue

            qso_date, qso_time = query.value(self.__sql_cols__.index('date_time')).split()

            record = {'QSO_DATE': qso_date.replace('-', ''),
                      'TIME_ON': qso_time.replace(':', '')}

            if query.value(self.__sql_cols__.index('call_sign')):
                record['CALL'] = query.value(self.__sql_cols__.index('call_sign'))
            if query.value(self.__sql_cols__.index('name')):
                record['NAME'] = self.replaceUmlautsLigatures(query.value(self.__sql_cols__.index('name')))
                record['NAME_INTL'] = query.value(self.__sql_cols__.index('name'))
            if query.value(self.__sql_cols__.index('qth')):
                record['QTH'] = self.replaceUmlautsLigatures(query.value(self.__sql_cols__.index('qth')))
                record['QTH_INTL'] = query.value(self.__sql_cols__.index('qth'))
            if query.value(self.__sql_cols__.index('locator')):
                record['GRIDSQUARE'] = query.value(self.__sql_cols__.index('locator'))
            if query.value(self.__sql_cols__.index('rst_sent')):
                record['RST_SENT'] = query.value(self.__sql_cols__.index('rst_sent'))
            if query.value(self.__sql_cols__.index('rst_rcvd')):
                record['RST_RCVD'] = query.value(self.__sql_cols__.index('rst_rcvd'))
            if band:
                if band == '11m':
                    if is_adx:
                        record['APP'] = [{'@PROGRAMID': 'DRAGONLOG',
                                          '@FIELDNAME': 'CBQSO',
                                          '@TYPE': 'B',
                                          '$': 'Y'}, ]
                        if query.value(self.__sql_cols__.index("channel")):
                            record['APP'].append({'@PROGRAMID': 'DRAGONLOG',
                                                  '@FIELDNAME': 'CHANNEL',
                                                  '@TYPE': 'N',
                                                  '$': str(query.value(self.__sql_cols__.index("channel")))})
                    else:
                        record['BAND'] = band.upper()
                        record['APP_DRAGONLOG_CBQSO'] = 'Y'
                        if query.value(self.__sql_cols__.index("channel")):
                            record['APP_DRAGONLOG_CHANNEL'] = query.value(self.__sql_cols__.index("channel"))
                else:
                    record['BAND'] = band.upper()
            if query.value(self.__sql_cols__.index('mode')):
                record['MODE'] = query.value(self.__sql_cols__.index('mode'))
            if query.value(self.__sql_cols__.index("freq")):
                record['FREQ'] = f'{query.value(self.__sql_cols__.index("freq")) / 1000:0.3f}'
            if query.value(self.__sql_cols__.index('power')):
                record['TX_PWR'] = query.value(self.__sql_cols__.index('power'))
            if query.value(self.__sql_cols__.index('own_callsign')):
                record['STATION_CALLSIGN'] = query.value(self.__sql_cols__.index('own_callsign'))
            if query.value(self.__sql_cols__.index('own_name')):
                record['MY_NAME'] = self.replaceUmlautsLigatures(query.value(self.__sql_cols__.index('own_name')))
                record['MY_NAME_INTL'] = query.value(self.__sql_cols__.index('own_name'))
            if query.value(self.__sql_cols__.index('own_qth')):
                record['MY_CITY'] = self.replaceUmlautsLigatures(query.value(self.__sql_cols__.index('own_qth')))
                record['MY_CITY_INTL'] = query.value(self.__sql_cols__.index('own_qth'))
            if query.value(self.__sql_cols__.index('own_locator')):
                record['MY_GRIDSQUARE'] = query.value(self.__sql_cols__.index('own_locator'))
            if query.value(self.__sql_cols__.index('radio')):
                record['MY_RIG'] = self.replaceUmlautsLigatures(query.value(self.__sql_cols__.index('radio')))
                record['MY_RIG_INTL'] = query.value(self.__sql_cols__.index('radio'))
            if query.value(self.__sql_cols__.index('antenna')):
                record['MY_ANTENNA'] = self.replaceUmlautsLigatures(query.value(self.__sql_cols__.index('antenna')))
                record['MY_ANTENNA_INTL'] = query.value(self.__sql_cols__.index('antenna'))
            if query.value(self.__sql_cols__.index('dist')):
                record['DISTANCE'] = query.value(self.__sql_cols__.index('dist'))
            if query.value(self.__sql_cols__.index('remarks')):
                record['NOTES'] = self.replaceUmlautsLigatures(
                    query.value(self.__sql_cols__.index('remarks')).replace('\n', '\r\n'))
                record['NOTES_INTL'] = query.value(self.__sql_cols__.index('remarks')).replace('\n', '\r\n')

            records.append(record)

        if os.path.splitext(file)[-1] == '.adx':
            doc['RECORDS'] = {'RECORD': records}
            ElementTree(self.adx_export_schema.encode(doc)).write(file, xml_declaration=True, encoding='utf-8')
        else:
            doc['RECORDS'] = records
            with open(file, 'w', encoding='ascii') as af:
                af.write(adif_file.dict2adi(doc, 'ADIF Export by DragonLog'))

        print(f'Saved "{file}"')

    def logImport(self):
        imp_formats = {
            # self.tr('Excel-File (*.xlsx)'): self.logImportExcel,
            self.tr('CSV-File (*.csv)'): self.logImportCSV,
            self.tr('ADIF 3 (*.adx *.adi *.adif)'): self.logImportADIF,
        }

        res = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr('Select import file'),
            self.settings.value('lastImportDir', os.path.abspath(os.curdir)),
            ';;'.join(imp_formats.keys())
        )

        if res[0]:
            imp_formats[res[1]](res[0])

            self.settings.setValue('lastImportDir', os.path.dirname(res[0]))

    def logImportCSV(self, file):
        print('Importing from CSV...')

        with open(file, newline='', encoding='utf-8') as cf:
            reader = csv.reader(cf, delimiter=';', dialect=csv.excel)

            ln = 0
            for row in reader:
                ln += 1
                if ln == 1:
                    continue

                if len(row) >= len(self.__sql_cols__) and row[1]:
                    query = QtSql.QSqlQuery(self.__db_con__)
                    query.prepare(self.__db_insert_stmnt__)

                    for i, val in enumerate(row[1:]):
                        query.bindValue(i, val)
                    query.exec()
                    if query.lastError().text():
                        QtWidgets.QMessageBox.warning(
                            self,
                            self.tr('Log import CSV'),
                            f'Row {ln} import error ("{query.lastError().text()}").\nSkipped row.'
                        )
                else:
                    QtWidgets.QMessageBox.warning(
                        self,
                        self.tr('Log import CSV'),
                        f'Row {ln} has too few columns.\nSkipped row.'
                    )

        self.__db_con__.commit()
        self.QSOTableView.model().select()
        self.QSOTableView.resizeColumnsToContents()
        print(f'Imported {ln - 1} QSOs from "{file}"')

    def logImportADIF(self, file):
        print('Importing from ADIF...')

        is_adx: bool = os.path.splitext(file)[-1] == '.adx'

        if is_adx:
            records: list = self.adx_import_schema.to_dict(file, decimal_type=str)['RECORDS']['RECORD']
        else:
            with open(file, encoding='ascii') as af:
                adi = af.read()
            records: list = adif_file.adi2dict(adi)['RECORDS']

        # align access to adx and adi records
        def rec_data(rec, param):
            return rec[param][0] if is_adx else rec[param]

        imported = 0
        for i, r in enumerate(records, 1):
            values = [''] * (len(self.__sql_cols__) - 1)

            if 'QSO_DATE' not in r or 'TIME_ON' not in r:
                QtWidgets.QMessageBox.warning(
                    self,
                    self.tr('Log import ADIF'),
                    f'QSO date/time missing in record {i}.\nSkipped record.'
                )
                continue

            date = rec_data(r, 'QSO_DATE')
            timex = rec_data(r, 'TIME_ON')
            time = f'{timex[:2]}:{timex[2:4]}' if len(timex) == 4 else f'{timex[:2]}:{timex[2:4]}:{timex[4:6]}'
            values[0] = f'{date[:4]}-{date[4:6]}-{date[6:8]} {time}'

            for p in r:
                match p:
                    case 'QSO_DATE' | 'TIME_ON':  # Skip date and time as they are already imported above
                        continue
                    case 'BAND':
                        values[self.__adx_cols__.index(p)] = rec_data(r, p).lower()
                    case 'FREQ':
                        values[self.__adx_cols__.index(p)] = str(float(rec_data(r, p)) * 1000)
                    case 'APP':  # Only for ADX as ADI App fields are recognised the standard way
                        for af in r[p]:
                            af_param = f'APP_{af["@PROGRAMID"].upper()}_{af["@FIELDNAME"].upper()}'
                            if af_param in self.__adx_cols__:
                                values[self.__adx_cols__.index(af_param)] = af['$']
                            elif af_param == 'APP_DRAGONLOG_CBQSO' and af['$'] == 'Y':
                                values[self.__adx_cols__.index('BAND')] = '11m'
                    case p if p in self.__adx_cols__:
                        values[self.__adx_cols__.index(p)] = rec_data(r, p)
                    case p if p + '_INTL' not in r:  # Take non *_INTL only if no suiting *_INTL are in import
                        values[self.__adx_cols__.index(p + '_INTL')] = rec_data(r, p)

            query = QtSql.QSqlQuery(self.__db_con__)
            query.prepare(self.__db_insert_stmnt__)

            for j, val in enumerate(values):
                query.bindValue(j, val)
            query.exec()
            if query.lastError().text():
                QtWidgets.QMessageBox.warning(
                    self,
                    self.tr('Log import ADIF'),
                    f'Record {i} import error ("{query.lastError().text()}").\nSkipped record.'
                )

            imported = i

        self.__db_con__.commit()
        self.QSOTableView.model().select()
        self.QSOTableView.resizeColumnsToContents()

        print(f'Imported {imported} QSOs from "{file}"')

    # noinspection PyPep8Naming
    def showHelp(self):
        if not self.help_dialog:
            with open(os.path.join(self.app_path, 'README.md')) as hf:
                help_text = hf.read()

            self.help_dialog = QtWidgets.QDialog(self)
            self.help_dialog.setWindowTitle(f'{self.tr(__prog_name__)} - {self.tr("Help")}')
            self.help_dialog.resize(500, 500)
            verticalLayout = QtWidgets.QVBoxLayout(self.help_dialog)
            scrollArea = QtWidgets.QScrollArea(self.help_dialog)
            scrollArea.setWidgetResizable(True)
            scrollAreaWidgetContents = QtWidgets.QWidget()
            verticalLayout_2 = QtWidgets.QVBoxLayout(scrollAreaWidgetContents)
            helpLabel = QtWidgets.QLabel(scrollAreaWidgetContents)
            helpLabel.setTextFormat(QtCore.Qt.TextFormat.MarkdownText)
            helpLabel.setWordWrap(True)
            helpLabel.setOpenExternalLinks(True)
            verticalLayout_2.addWidget(helpLabel)
            scrollArea.setWidget(scrollAreaWidgetContents)
            verticalLayout.addWidget(scrollArea)
            horizontalLayout = QtWidgets.QHBoxLayout()
            horizontalSpacer = QtWidgets.QSpacerItem(40, 20,
                                                     QtWidgets.QSizePolicy.Policy.Expanding,
                                                     QtWidgets.QSizePolicy.Policy.Minimum)
            horizontalLayout.addItem(horizontalSpacer)
            pushButton = QtWidgets.QPushButton(self.help_dialog)
            pushButton.setText(self.tr('Ok'))
            horizontalLayout.addWidget(pushButton)
            verticalLayout.addLayout(horizontalLayout)
            # noinspection PyUnresolvedReferences
            pushButton.clicked.connect(self.help_dialog.accept)

            helpLabel.setText(help_text)

        self.help_dialog.show()

    def showAbout(self):
        cr = sys.copyright.replace('\n\n', '\n')

        QtWidgets.QMessageBox.about(
            self,
            f'{__prog_name__} - {self.tr("About")}',
            f'{self.tr("Version")}: {__version__}\n'
            f'{self.tr("Author")}: {__author_name__} <{__author_email__}>\n{__copyright__}'
            f'\n\nPython {sys.version.split()[0]}: {cr}'
            f'\n\nOpenPyXL {openpyxl.__version__}: Copyright (c) 2010 openpyxl'
            '\nmaidenhead: Copyright (c) 2018 Michael Hirsch, Ph.D.' +
            f'\nxmlschema {xmlschema.__version__}: Copyright (c), 2016-2022, '
            f'SISSA (Scuola Internazionale Superiore di Studi Avanzati)' +
            f'\nPyADIF-File {adif_file.__version_str__}: {adif_file.__copyright__}' +
            '\n\nIcons: Crystal Project, Copyright (c) 2006-2007 Everaldo Coelho'
            '\nDragon icon by Icons8 https://icons8.com'
        )

    def showAboutQt(self):
        QtWidgets.QMessageBox.aboutQt(self, __prog_name__ + ' - ' + self.tr('About Qt'))

    def closeEvent(self, e):
        print(f'Quiting {__prog_name__}...')
        self.__db_con__.close()
        self.settings_form.ctrlRigctld(False)
        e.accept()


def main():
    app = QtWidgets.QApplication(sys.argv)
    translator = QtCore.QTranslator(app)
    translator.load(os.path.abspath(os.path.dirname(sys.argv[0])) + '/DragonLog_' + QtCore.QLocale.system().name())
    app.installTranslator(translator)

    file = None
    args = app.arguments()
    if len(args) > 1:
        file = args[1]
        if not (os.path.isfile(file) and os.path.splitext(file)[-1] in ('.qlog', '.sqlite')):
            file = None

    app_path = os.path.dirname(args[0])

    QtCore.QDir.addSearchPath('icons', app_path + '/icons')
    QtCore.QDir.addSearchPath('data', app_path + '/data')

    dl = DragonLog(file, app_path)
    dl.show()

    sys.exit(app.exec())


# Get old behaviour on printing a traceback on exceptions
def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


if __name__ == '__main__':
    sys.excepthook = except_hook

    main()
